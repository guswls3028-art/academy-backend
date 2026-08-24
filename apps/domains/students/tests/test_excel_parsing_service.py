from pathlib import Path

import pytest
from openpyxl import Workbook

from academy.application.services.excel_parsing_service import (
    ExcelValidationError,
    parse_student_excel_file,
)


def _write_student_excel(path: Path, *, name: str, parent_phone: str = "01031217466") -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["이름", "학부모전화번호", "학생전화번호", "학교", "학년", "성별", "메모"])
    ws.append([name, parent_phone, "", "테스트고", 1, "M", "parser-regression"])
    wb.save(path)


def test_parse_student_excel_allows_long_name_when_valid_phone_exists(tmp_path):
    path = tmp_path / "students.xlsx"
    name = "E2E-ALIM-0523115013학생"
    assert len(name) > 20
    _write_student_excel(path, name=name)

    rows, _lecture_title = parse_student_excel_file(str(path))

    assert len(rows) == 1
    assert rows[0]["name"] == name
    assert rows[0]["parent_phone"] == "01031217466"


def test_parse_student_excel_remains_strict_without_partial_error_collector(tmp_path):
    path = tmp_path / "invalid-students.xlsx"
    _write_student_excel(path, name="오류학생", parent_phone="번호오류")

    with pytest.raises(ExcelValidationError) as exc_info:
        parse_student_excel_file(str(path))

    assert exc_info.value.errors[0]["row"] == 2


def test_parse_student_excel_keeps_student_and_guardian_contacts_separate(tmp_path):
    path = tmp_path / "contact-columns.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["학생 성명", "학생 연락처", "보호자 휴대전화"])
    ws.append(["연락처학생", "01012345678", "01087654321"])
    wb.save(path)

    rows, _lecture_title = parse_student_excel_file(str(path))

    assert len(rows) == 1
    assert rows[0]["phone"] == "01012345678"
    assert rows[0]["parent_phone"] == "01087654321"


def test_parse_student_excel_treats_matching_contacts_as_parent_only(tmp_path):
    path = tmp_path / "shared-contact.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["학생 성명", "학생 연락처", "보호자 휴대전화"])
    ws.append(["무휴대폰학생", "010-8765-4321", "010-8765-4321"])
    wb.save(path)

    rows, _lecture_title = parse_student_excel_file(str(path))

    assert len(rows) == 1
    assert rows[0]["phone"] is None
    assert rows[0]["studentPhone"] is None
    assert rows[0]["parent_phone"] == "01087654321"
    assert rows[0]["uses_identifier"] is True


def test_parse_student_excel_selects_data_sheet_behind_cover_sheet(tmp_path):
    path = tmp_path / "cover-and-roster.xlsx"
    wb = Workbook()
    cover = wb.active
    cover.title = "안내"
    cover.append(["학생 등록 안내"])
    cover.append(["작성 후 명단 시트를 업로드하세요."])
    roster = wb.create_sheet("학생 명단")
    roster.append(["Student Name", "Guardian Phone", "School / Grade"])
    roster.append(["멀티시트학생", "010-7777-8888", "테스트고(2)"])
    wb.save(path)

    rows, _lecture_title = parse_student_excel_file(str(path))

    assert len(rows) == 1
    assert rows[0]["name"] == "멀티시트학생"
    assert rows[0]["parent_phone"] == "01077778888"
    assert rows[0]["school"] == "테스트고"
    assert rows[0]["grade"] == "2"


def test_parse_student_excel_accepts_one_row_with_unambiguous_phone_data(tmp_path):
    path = tmp_path / "single-row-custom-headers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["수강생", "비상번호"])
    ws.append(["단일학생", "010-9999-0000"])
    wb.save(path)

    rows, _lecture_title = parse_student_excel_file(str(path))

    assert len(rows) == 1
    assert rows[0]["name"] == "단일학생"
    assert rows[0]["parent_phone"] == "01099990000"


def test_parse_student_excel_fails_closed_for_student_only_contact(tmp_path):
    path = tmp_path / "student-contact-only.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["이름", "학생 연락처"])
    ws.append(["보호자없음", "01011112222"])
    wb.save(path)

    with pytest.raises(ExcelValidationError, match="학부모 전화번호"):
        parse_student_excel_file(str(path))


def test_parse_student_excel_fails_closed_for_ambiguous_roster_sheets(tmp_path):
    path = tmp_path / "ambiguous-rosters.xlsx"
    wb = Workbook()
    cover = wb.active
    cover.title = "안내"
    cover.append(["학생 등록 안내"])
    for title, phone in (("중2 명단", "010-1111-2222"), ("중3 명단", "010-3333-4444")):
        roster = wb.create_sheet(title)
        roster.append(["학생명", "보호자 연락처"])
        roster.append([f"{title}학생", phone])
    wb.save(path)

    with pytest.raises(ExcelValidationError, match="시트가 여러 개"):
        parse_student_excel_file(str(path))
