# apps/domains/attendance/utils/excel.py
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from academy.adapters.db.django import repositories_enrollment as enroll_repo


STATUS_LABEL_MAP = {
    "PRESENT": "현장",
    "LATE": "지각",
    "ONLINE": "영상",
    "SUPPLEMENT": "보강",
    "EARLY_LEAVE": "조퇴",
    "ABSENT": "결석",
    "RUNAWAY": "출튀",
    "MATERIAL": "자료",
    "INACTIVE": "부재",
    "SECESSION": "퇴원",
}

STATUS_FILL_MAP = {
    "PRESENT": "C6EFCE",
    "ABSENT": "FFC7CE",
    "LATE": "FFEB9C",
    "ONLINE": "BDD7EE",
    "SUPPLEMENT": "B7DEE8",
    "EARLY_LEAVE": "FFEB9C",
    "RUNAWAY": "FFC7CE",
    "MATERIAL": "D9D9D9",
    "INACTIVE": "F2F2F2",
    "SECESSION": "BDBCBC",
}


def build_attendance_excel(lecture):
    sessions = enroll_repo.get_sessions_for_lecture_ordered(lecture)
    enrollment_ids = list(enroll_repo.get_session_enrollment_enrollment_ids_by_lecture(lecture))
    tenant = getattr(lecture, "tenant", None) or getattr(lecture, "tenant_id", None)
    enrollments = enroll_repo.get_enrollments_by_ids_active(enrollment_ids, tenant)
    attendances = enroll_repo.get_attendances_for_lecture_by_lecture(lecture, enrollments)

    attendance_map = {
        (a.enrollment_id, a.session_id): a
        for a in attendances
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "출결"

    # Header
    header = ["학생명", "학생번호", "학부모번호"]
    for s in sessions:
        label = f"{s.order}차시"
        if s.date:
            label += f" ({s.date})"
        header.append(label)

    ws.append(header)

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.freeze_panes = "D2"

    for en in enrollments:
        row = [
            en.student.name,
            en.student.phone or "",
            en.student.parent_phone or "",
        ]

        for s in sessions:
            att = attendance_map.get((en.id, s.id))
            code = att.status if att else ""
            label = STATUS_LABEL_MAP.get(code, code)
            row.append(label)

        ws.append(row)
        r = ws.max_row

        for idx, s in enumerate(sessions, start=4):
            att = attendance_map.get((en.id, s.id))
            if att and att.status in STATUS_FILL_MAP:
                ws.cell(row=r, column=idx).fill = PatternFill(
                    start_color=STATUS_FILL_MAP[att.status],
                    end_color=STATUS_FILL_MAP[att.status],
                    fill_type="solid",
                )
            ws.cell(row=r, column=idx).alignment = center

    filename = f"출결_{lecture.title}_{lecture.id}.xlsx"
    return wb, filename
