# apps/domains/attendance/views.py
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.http import HttpResponse

from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import action
from rest_framework.filters import SearchFilter
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.response import Response
from rest_framework import status

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Attendance
from .serializers import (
    AttendanceSerializer,
    AttendanceMatrixStudentSerializer,
)
from .filters import AttendanceFilter

from apps.domains.lectures.models import Lecture, Session
from apps.domains.enrollment.models import Enrollment, SessionEnrollment


STATUS_LABEL_MAP = {
    "PRESENT": "현장",
    "LATE": "지각",
    "ONLINE": "영상",
    "SUPPLEMENT": "보강",
    "EARLY_LEAVE": "조퇴",
    "ABSENT": "결석",
    "RUNAWAY": "출튀",
    "MATERIAL": "자료",
    "INACTIVE": "부재",
    "SECESSION": "퇴원",
}

# 엑셀 색상(원하면 더 추가 가능)
STATUS_FILL_MAP = {
    "PRESENT": "C6EFCE",   # 연두
    "ABSENT": "FFC7CE",    # 연빨강
    "LATE": "FFEB9C",      # 연노랑
    "ONLINE": "BDD7EE",    # 연파랑
}


class AttendanceViewSet(ModelViewSet):
    queryset = Attendance.objects.all().select_related(
        "session",
        "enrollment",
        "enrollment__student",
    )
    serializer_class = AttendanceSerializer

    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_class = AttendanceFilter
    search_fields = ["enrollment__student__name"]

    # =========================================================
    # 1) 세션 기준 학생 등록 → 강의 자동 등록 → 출결 생성
    # =========================================================
    @transaction.atomic
    @action(detail=False, methods=["post"])
    def bulk_create(self, request):
        session_id = request.data.get("session")
        student_ids = request.data.get("students", [])

        if not session_id or not isinstance(student_ids, list):
            return Response(
                {"detail": "session, students(list)는 필수입니다"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        session = get_object_or_404(Session, id=session_id)
        created = []

        for sid in student_ids:
            enrollment, _ = Enrollment.objects.get_or_create(
                student_id=sid,
                lecture=session.lecture,
                defaults={"status": "ACTIVE"},
            )

            SessionEnrollment.objects.get_or_create(
                enrollment=enrollment,
                session=session,
            )

            attendance, _ = Attendance.objects.get_or_create(
                enrollment=enrollment,
                session=session,
                defaults={"status": "PRESENT"},
            )

            created.append(attendance)

        return Response(
            AttendanceSerializer(created, many=True).data,
            status=status.HTTP_201_CREATED,
        )

    # =========================================================
    # 2) 강의 × 차시 출결 매트릭스 API (GET)
    #    GET /lectures/attendance/matrix/?lecture={id}
    # =========================================================
    @action(detail=False, methods=["get"], url_path="matrix")
    def matrix(self, request):
        lecture_id = request.query_params.get("lecture")

        if not lecture_id:
            return Response(
                {"detail": "lecture 파라미터는 필수입니다"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        lecture = get_object_or_404(Lecture, id=lecture_id)

        sessions = Session.objects.filter(
            lecture=lecture
        ).order_by("order", "id")

        enrollments = Enrollment.objects.filter(
            lecture=lecture,
            status="ACTIVE",
        ).select_related("student").order_by("student__name", "id")

        attendances = Attendance.objects.filter(
            session__lecture=lecture,
            enrollment__in=enrollments,
        ).select_related("session", "enrollment")

        attendance_map = {
            (a.enrollment_id, a.session_id): a
            for a in attendances
        }

        students_payload = []

        for en in enrollments:
            row = {
                "student_id": en.student.id,
                "name": en.student.name,
                "phone": en.student.phone,
                "parent_phone": en.student.parent_phone,
                "attendance": {},
            }

            for s in sessions:
                att = attendance_map.get((en.id, s.id))
                if att:
                    row["attendance"][str(s.id)] = {
                        "attendance_id": att.id,
                        "status": att.status,
                    }
                else:
                    # (선택) 빈 셀도 프론트에서 안전하게 다루고 싶으면 주석 해제
                    # row["attendance"][str(s.id)] = None
                    pass

            students_payload.append(row)

        return Response(
            {
                "lecture": {
                    "id": lecture.id,
                    "title": lecture.title,
                },
                "sessions": [
                    {
                        "id": s.id,
                        "order": s.order,
                        "date": s.date,
                        "title": getattr(s, "title", None),
                    }
                    for s in sessions
                ],
                "students": AttendanceMatrixStudentSerializer(
                    students_payload, many=True
                ).data,
            },
            status=status.HTTP_200_OK,
        )

    # =========================================================
    # 3) 엑셀 다운로드 API
    #    GET /lectures/attendance/excel/?lecture={id}
    # =========================================================
    @action(detail=False, methods=["get"], url_path="excel")
    def excel(self, request):
        lecture_id = request.query_params.get("lecture")
        if not lecture_id:
            return Response(
                {"detail": "lecture 파라미터는 필수입니다"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        lecture = get_object_or_404(Lecture, id=lecture_id)

        sessions = Session.objects.filter(lecture=lecture).order_by("order", "id")
        enrollments = Enrollment.objects.filter(
            lecture=lecture, status="ACTIVE"
        ).select_related("student").order_by("student__name", "id")

        attendances = Attendance.objects.filter(
            session__lecture=lecture,
            enrollment__in=enrollments,
        ).select_related("session", "enrollment", "enrollment__student")

        attendance_map = {(a.enrollment_id, a.session_id): a for a in attendances}

        wb = Workbook()
        ws = wb.active
        ws.title = "출결"

        # ----- Header
        header = ["학생명", "학생번호", "학부모번호"]
        for s in sessions:
            # 예: "1차시 (2025-01-01)"
            label = f"{s.order}차시"
            if getattr(s, "date", None):
                label += f" ({str(s.date)})"
            header.append(label)

        ws.append(header)

        # 스타일
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = center
            ws.column_dimensions[get_column_letter(col)].width = 16

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16

        ws.freeze_panes = "D2"  # A~C 고정 + 헤더 고정 느낌

        # ----- Rows
        for en in enrollments:
            st = en.student
            row = [
                st.name,
                st.phone or "",
                st.parent_phone or "",
            ]

            for s in sessions:
                att = attendance_map.get((en.id, s.id))
                code = att.status if att else ""
                label = STATUS_LABEL_MAP.get(code, code)
                row.append(label)

            ws.append(row)

            # 방금 추가된 row index
            r = ws.max_row

            # 셀 스타일 + 색
            for c in range(1, len(header) + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = center

            # 출결 영역만 색칠(D~)
            for idx, s in enumerate(sessions, start=4):
                label = ws.cell(row=r, column=idx).value
                # label -> code 역변환 대신, map에서 code를 다시 뽑는 방식
                att = attendance_map.get((en.id, s.id))
                code = att.status if att else None
                if code and code in STATUS_FILL_MAP:
                    ws.cell(row=r, column=idx).fill = PatternFill(
                        start_color=STATUS_FILL_MAP[code],
                        end_color=STATUS_FILL_MAP[code],
                        fill_type="solid",
                    )

        filename = f"출결_{lecture.title}_{lecture.id}.xlsx"
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp

    # =========================================================
    # 3️⃣ 엑셀 다운로드 API
    # =========================================================
    @action(detail=False, methods=["get"], url_path="excel")
    def excel(self, request):
        lecture_id = request.query_params.get("lecture")
        if not lecture_id:
            return Response(
                {"detail": "lecture 파라미터는 필수입니다"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        lecture = Lecture.objects.get(id=lecture_id)

        sessions = Session.objects.filter(
            lecture=lecture
        ).order_by("order")

        enrollments = Enrollment.objects.filter(
            lecture=lecture,
            status="ACTIVE",
        ).select_related("student")

        attendances = Attendance.objects.filter(
            session__lecture=lecture,
            enrollment__in=enrollments,
        )

        attendance_map = {
            (a.enrollment_id, a.session_id): a.status
            for a in attendances
        }

        # -----------------------------
        # Workbook
        # -----------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = "출결 현황"

        # 헤더
        headers = ["이름", "학생 전화", "학부모 전화"]
        for s in sessions:
            headers.append(f"{s.order}차시")

        ws.append(headers)

        # 컬러맵
        STATUS_FILL = {
            "PRESENT": PatternFill("solid", fgColor="C6EFCE"),   # 초록
            "LATE": PatternFill("solid", fgColor="FFEB9C"),      # 노랑
            "ONLINE": PatternFill("solid", fgColor="BDD7EE"),    # 파랑
            "SUPPLEMENT": PatternFill("solid", fgColor="E4DFEC"),# 보라
            "EARLY_LEAVE": PatternFill("solid", fgColor="FCE4D6"),# 주황
            "ABSENT": PatternFill("solid", fgColor="FFC7CE"),    # 빨강
            "RUNAWAY": PatternFill("solid", fgColor="F4B6C2"),   # 진빨강
            "MATERIAL": PatternFill("solid", fgColor="D9E1F2"),  # 남보라
            "INACTIVE": PatternFill("solid", fgColor="E7E6E6"),  # 회색
            "SECESSION": PatternFill("solid", fgColor="D0CECE"), # 탈퇴
        }

        # 데이터
        for en in enrollments:
            row = [
                en.student.name,
                en.student.phone,
                en.student.parent_phone,
            ]

            for s in sessions:
                status = attendance_map.get((en.id, s.id), "")
                row.append(status)

            ws.append(row)

            # 상태 셀 컬러
            r_idx = ws.max_row
            for c_idx, s in enumerate(sessions, start=4):
                cell = ws.cell(row=r_idx, column=c_idx)
                fill = STATUS_FILL.get(cell.value)
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(horizontal="center")

        # 컬럼 자동 너비
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

        # Response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{lecture.title}_출결현황.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response