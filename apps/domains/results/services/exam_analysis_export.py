from __future__ import annotations

import io
import math
from dataclasses import dataclass
from statistics import median, pstdev
from typing import Any

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.domains.results.models import ExamAttempt, ResultFact
from apps.domains.results.services.exam_result_excel_import import (
    ExamResultWorkbookError,
)
from apps.domains.results.services.question_stats_service import QuestionStatsService
from apps.domains.results.utils.exam_achievement import compute_exam_achievement_bulk
from apps.domains.results.utils.initial_exam_score import (
    load_initial_exam_scores,
    project_initial_exam_score,
)
from apps.domains.results.utils.ranking import compute_exam_rankings
from apps.domains.results.utils.result_queries import latest_results_per_enrollment
from apps.domains.results.utils.session_exam import get_primary_session_for_exam
from apps.support.results.exam_result_excel_import_dependencies import (
    get_result_import_candidates,
    get_result_import_questions,
)
from apps.support.omr.sheet_resolver import resolve_omr_sheet_for_exam


_NAVY = "172554"
_BLUE = "2563EB"
_BLUE_SOFT = "E8F0FF"
_SLATE = "334155"
_SLATE_SOFT = "F1F5F9"
_MUTED = "64748B"
_BORDER = "CBD5E1"
_GREEN = "15803D"
_GREEN_SOFT = "DCFCE7"
_AMBER = "B45309"
_AMBER_SOFT = "FEF3C7"
_RED = "B91C1C"
_RED_SOFT = "FEE2E2"
_WHITE = "FFFFFF"

_THIN_BORDER = Border(
    left=Side(style="thin", color=_BORDER),
    right=Side(style="thin", color=_BORDER),
    top=Side(style="thin", color=_BORDER),
    bottom=Side(style="thin", color=_BORDER),
)


@dataclass(frozen=True)
class ExamBriefing:
    direction: str
    direction_detail: str
    cut_review: str
    cut_review_detail: str
    next_action: str
    next_action_detail: str


def _safe_excel_text(value: Any) -> str:
    text = str(value or "")
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _mean(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def _top_ten_mean(values: list[float]) -> float:
    if not values:
        return 0.0
    count = max(1, math.ceil(len(values) * 0.1))
    return _mean(sorted(values, reverse=True)[:count])


def _score_rate(score: float, max_score: float) -> float:
    return max(0.0, min(100.0, (score / max_score) * 100.0)) if max_score > 0 else 0.0


def _briefing(
    *,
    scored_count: int,
    pass_rate: float,
    std_rate: float,
    question_stats: list[dict[str, Any]],
    fail_count: int,
    pass_score: float,
    has_pass_criterion: bool,
) -> ExamBriefing:
    weak = [row for row in question_stats if float(row.get("accuracy") or 0.0) < 0.5]
    critical = [row for row in question_stats if float(row.get("accuracy") or 0.0) < 0.3]

    if scored_count < 5:
        direction = "표본 확인 후 판단"
        direction_detail = "응시 인원이 5명 미만입니다. 개인 결과를 함께 보고 수업 방향을 확정하세요."
    elif (has_pass_criterion and pass_rate < 0.4) or len(critical) >= 2:
        direction = "전체 재설명 우선"
        direction_detail = "미달 비율 또는 최저 정답률 문항이 높아 공통 개념부터 다시 설명하는 편이 안전합니다."
    elif std_rate >= 20.0:
        direction = "수준별 보충 권장"
        direction_detail = "점수 편차가 커서 공통 설명 뒤 상·중·하 난이도로 재풀이를 나누는 편이 효율적입니다."
    elif weak:
        direction = "취약 문항 재풀이"
        direction_detail = "전체 흐름은 유지하고 정답률 50% 미만 문항을 중심으로 보충하세요."
    else:
        direction = "현재 수업 흐름 유지"
        direction_detail = "합격률과 문항별 정답률이 안정적입니다. 오답 확인 중심으로 마무리할 수 있습니다."

    fail_rate = 1.0 - pass_rate if scored_count else 0.0
    if not has_pass_criterion:
        cut_review = "합격 기준 설정 필요"
        cut_review_detail = "합격 컷이 설정되지 않아 합격·미달 인원을 계산하지 않았습니다. 시험 설정에서 기준 점수를 먼저 확인하세요."
    elif scored_count < 5:
        cut_review = "컷 판단 보류"
        cut_review_detail = f"현재 {pass_score:g}점 기준을 유지하고 표본이 쌓인 뒤 검토하세요."
    elif fail_rate >= 0.6:
        cut_review = "난이도·문항 오류 먼저 확인"
        cut_review_detail = f"현재 컷 {pass_score:g}점에서 {fail_count}명이 미달입니다. 컷 변경 전 시험 난이도와 문항 오류를 확인하세요."
    elif fail_rate <= 0.1:
        cut_review = "현재 컷 유지 또는 상향 검토"
        cut_review_detail = f"현재 컷 {pass_score:g}점의 미달 비율이 10% 이하입니다. 수업 목표에 따라 다음 시험부터 조정할 수 있습니다."
    else:
        cut_review = "현재 컷으로 운영 가능"
        cut_review_detail = f"현재 컷 {pass_score:g}점에서 {fail_count}명이 미달입니다. 이번 시험은 현 기준으로 보충 대상을 운영하세요."

    if critical:
        numbers = ", ".join(f"{int(row['question_number'])}번" for row in critical[:3])
        next_action = f"{numbers} 공통 해설"
        next_action_detail = "정답률 30% 미만 문항입니다. 개념 확인 → 대표 풀이 → 유사 문항 순서로 보충하세요."
    elif weak:
        numbers = ", ".join(f"{int(row['question_number'])}번" for row in weak[:3])
        next_action = f"{numbers} 재풀이"
        next_action_detail = "정답률이 낮은 순서입니다. 학생별 오답표와 함께 대상자를 나누어 전달하세요."
    else:
        next_action = "학생별 오답 확인"
        next_action_detail = "공통 취약 문항이 뚜렷하지 않습니다. 개인 오답과 복습 지정 문항을 확인하세요."

    return ExamBriefing(
        direction=direction,
        direction_detail=direction_detail,
        cut_review=cut_review,
        cut_review_detail=cut_review_detail,
        next_action=next_action,
        next_action_detail=next_action_detail,
    )


def _question_action(accuracy: float) -> tuple[str, str]:
    if accuracy < 0.3:
        return "최우선", "공통 개념 재설명"
    if accuracy < 0.5:
        return "우선", "풀이 시범 후 재시험"
    if accuracy < 0.7:
        return "보충", "유사 문항 추가"
    return "확인", "개별 오답 확인"


def _set_title(sheet, title: str, subtitle: str, *, last_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=last_column)
    cell = sheet.cell(1, 1, _safe_excel_text(title))
    cell.font = Font(size=19, bold=True, color=_WHITE)
    cell.fill = PatternFill("solid", fgColor=_NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 25
    sheet.row_dimensions[2].height = 18
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_column)
    cell = sheet.cell(3, 1, _safe_excel_text(subtitle))
    cell.font = Font(size=10, color=_MUTED)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[3].height = 22


def _section_heading(sheet, row: int, title: str, *, last_column: int) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)
    cell = sheet.cell(row, 1, title)
    cell.font = Font(size=12, bold=True, color=_NAVY)
    cell.fill = PatternFill("solid", fgColor=_BLUE_SOFT)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 23


def _table_header(sheet, row: int, headers: list[str]) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row, column, header)
        cell.font = Font(size=10, bold=True, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor=_SLATE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    sheet.row_dimensions[row].height = 25


def _style_data_table(sheet, *, start_row: int, end_row: int, end_column: int) -> None:
    for row in sheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=1,
        max_col=end_column,
    ):
        for cell in row:
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row % 2 == 0 and cell.fill.fill_type is None:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")


def _finalize_sheet(sheet, widths: list[float], *, freeze: str | None = None) -> None:
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.sheet_view.showGridLines = False
    if freeze:
        sheet.freeze_panes = freeze
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.outlinePr.summaryBelow = True


def build_exam_analysis_export(*, exam: Any, tenant: Any) -> bytes:
    """Build a teacher-facing workbook from the current representative exam state."""
    candidates = get_result_import_candidates(exam_id=int(exam.id), tenant=tenant)
    if not candidates:
        raise ExamResultWorkbookError("이 시험에 등록된 학생이 없습니다.")

    try:
        sheet = resolve_omr_sheet_for_exam(
            tenant=tenant,
            exam_id=int(exam.id),
            requested_sheet_id=None,
        )
    except ValueError:
        sheet = None
    questions = get_result_import_questions(sheet=sheet) if sheet is not None else []
    question_by_id = {int(question.question_id): question for question in questions}

    candidate_by_id = {int(candidate.enrollment_id): candidate for candidate in candidates}
    results = list(
        latest_results_per_enrollment(target_type="exam", target_id=int(exam.id))
        .filter(enrollment__tenant=tenant, enrollment_id__in=candidate_by_id)
        .select_related("attempt")
        .prefetch_related("items")
    )
    result_by_id = {int(result.enrollment_id): result for result in results}
    initial_states = load_initial_exam_scores(
        exam_ids=[int(exam.id)],
        enrollment_ids=candidate_by_id,
    )
    initial_scores = {
        int(result.enrollment_id): project_initial_exam_score(
            state=initial_states.get((int(exam.id), int(result.enrollment_id))),
            fallback_score=result.total_score,
            fallback_max_score=result.max_score,
            fallback_recorded_at=result.submitted_at or result.created_at,
        )
        for result in results
    }
    initial_attempt_ids = {
        int(state.attempt_id)
        for state in initial_states.values()
        if state.attempt_id is not None
    }
    initial_attempts = {
        int(attempt.id): attempt
        for attempt in ExamAttempt.objects.filter(id__in=initial_attempt_ids).only(
            "id", "status", "meta",
        )
    }
    fact_items_by_enrollment: dict[int, dict[int, Any]] = {}
    for fact in ResultFact.objects.filter(
        attempt_id__in=initial_attempt_ids,
        enrollment_id__in=candidate_by_id,
    ).order_by("-id"):
        fact_items_by_enrollment.setdefault(int(fact.enrollment_id), {}).setdefault(
            int(fact.question_id),
            fact,
        )

    def initial_items(result: Any | None) -> list[Any]:
        if result is None:
            return []
        enrollment_id = int(result.enrollment_id)
        state = initial_states.get((int(exam.id), enrollment_id))
        if state is None:
            return list(result.items.all()) if result.attempt_id is None else []
        if result.attempt_id == state.attempt_id:
            return list(result.items.all())
        return list(fact_items_by_enrollment.get(enrollment_id, {}).values())

    rankings = compute_exam_rankings(exam_id=int(exam.id), tenant=tenant)
    session = get_primary_session_for_exam(int(exam.id))
    pass_score = float(exam.pass_score or 0.0)
    has_pass_criterion = pass_score > 0
    achievement_map = compute_exam_achievement_bulk(
        items=[
            {
                "enrollment_id": int(result.enrollment_id),
                "exam_id": int(exam.id),
                "total_score": initial_scores[int(result.enrollment_id)].total_score,
                "pass_score": pass_score,
                "attempt_id": initial_scores[int(result.enrollment_id)].attempt_id,
                "session": session,
            }
            for result in results
        ],
        tenant=tenant,
    )

    def result_status(result: Any | None) -> str:
        if result is None:
            return "MISSING"
        achievement = achievement_map.get(
            (int(result.enrollment_id), int(exam.id)),
            {},
        )
        if achievement.get("meta_status") == "NOT_SUBMITTED":
            return "NOT_SUBMITTED"
        initial_score = initial_scores[int(result.enrollment_id)]
        attempt = initial_attempts.get(int(initial_score.attempt_id or 0))
        if attempt is None and initial_score.attempt_id is None:
            attempt = getattr(result, "attempt", None)
        attempt_status = str(getattr(attempt, "status", "") or "").lower()
        if attempt_status in {"pending", "grading"}:
            return "PROCESSING"
        if attempt_status == "failed":
            return "FAILED"
        if bool(achievement.get("is_provisional")):
            return "PARTIAL"
        return "DONE"

    def is_scored(result: Any | None) -> bool:
        return result_status(result) == "DONE"

    def analysis_score(result: Any) -> float:
        rank_info = rankings.get(int(result.enrollment_id), {})
        ranking_score = rank_info.get("ranking_score")
        return (
            float(ranking_score)
            if ranking_score is not None
            else float(initial_scores[int(result.enrollment_id)].total_score or 0.0)
        )

    scored_results = [result for result in results if is_scored(result)]
    scores = [analysis_score(result) for result in scored_results]
    scored_attempt_ids = [
        int(initial_scores[int(result.enrollment_id)].attempt_id)
        for result in scored_results
        if initial_scores[int(result.enrollment_id)].attempt_id is not None
    ]
    legacy_scored_enrollment_ids = [
        int(result.enrollment_id)
        for result in scored_results
        if initial_states.get((int(exam.id), int(result.enrollment_id))) is None
        and getattr(result, "attempt_id", None) is None
    ]
    question_stats = QuestionStatsService.per_question_stats(
        exam_id=int(exam.id),
        attempt_ids=scored_attempt_ids,
        legacy_enrollment_ids=legacy_scored_enrollment_ids,
    )
    question_stats.sort(
        key=lambda row: (
            float(row.get("accuracy") or 0.0),
            -int(row.get("attempts") or 0),
            int(row.get("question_number") or 0),
        )
    )
    max_score = float(exam.max_score or 100.0)
    pass_count = (
        sum(1 for score in scores if score >= pass_score)
        if has_pass_criterion
        else 0
    )
    fail_count = max(len(scores) - pass_count, 0) if has_pass_criterion else 0
    pass_rate = (
        pass_count / len(scores)
        if has_pass_criterion and scores
        else 0.0
    )
    score_rates = [_score_rate(score, max_score) for score in scores]
    std_rate = pstdev(score_rates) if score_rates else 0.0
    briefing = _briefing(
        scored_count=len(scores),
        pass_rate=pass_rate,
        std_rate=std_rate,
        question_stats=question_stats,
        fail_count=fail_count,
        pass_score=pass_score,
        has_pass_criterion=has_pass_criterion,
    )

    generated_at = timezone.localtime().strftime("%Y-%m-%d %H:%M")
    workbook = Workbook()
    workbook.properties.title = f"{exam.title} 수업 분석 리포트"
    workbook.properties.subject = "시험 결과 기반 수업 방향·보충·재시험 검토"
    workbook.properties.creator = "학원플러스"

    overview = workbook.active
    overview.title = "수업 브리핑"
    _set_title(
        overview,
        f"{exam.title} · 수업 분석 리포트",
        f"{getattr(exam, 'subject', '') or '과목 미지정'} · 사이트 최종 저장 기준 · {generated_at} · 제안은 자동으로 시험 정책을 변경하지 않습니다.",
        last_column=8,
    )
    _section_heading(overview, 5, "60초 수업 브리핑", last_column=8)
    cards = [
        (1, 3, "수업 방향", briefing.direction, briefing.direction_detail),
        (4, 6, "컷 검토", briefing.cut_review, briefing.cut_review_detail),
        (7, 8, "바로 할 일", briefing.next_action, briefing.next_action_detail),
    ]
    for start_column, end_column, label, value, detail in cards:
        overview.merge_cells(start_row=6, start_column=start_column, end_row=6, end_column=end_column)
        overview.merge_cells(start_row=7, start_column=start_column, end_row=7, end_column=end_column)
        overview.merge_cells(start_row=8, start_column=start_column, end_row=9, end_column=end_column)
        label_cell = overview.cell(6, start_column, label)
        label_cell.font = Font(size=9, bold=True, color=_BLUE)
        label_cell.fill = PatternFill("solid", fgColor=_SLATE_SOFT)
        label_cell.alignment = Alignment(vertical="center")
        value_cell = overview.cell(7, start_column, value)
        value_cell.font = Font(size=13, bold=True, color=_NAVY)
        value_cell.fill = PatternFill("solid", fgColor=_SLATE_SOFT)
        value_cell.alignment = Alignment(vertical="center", wrap_text=True)
        detail_cell = overview.cell(8, start_column, detail)
        detail_cell.font = Font(size=9, color=_MUTED)
        detail_cell.fill = PatternFill("solid", fgColor=_SLATE_SOFT)
        detail_cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row_index in range(6, 10):
            for column in range(start_column, end_column + 1):
                overview.cell(row_index, column).border = _THIN_BORDER
    overview.row_dimensions[8].height = 28
    overview.row_dimensions[9].height = 16

    _section_heading(overview, 11, "핵심 지표", last_column=8)
    metrics = [
        ("응시", f"{len(scores)}명"),
        ("미응시·미채점", f"{len(candidates) - len(scores)}명"),
        ("평균", f"{_mean(scores):.1f}/{max_score:g}"),
        ("상위 10% 평균", f"{_top_ten_mean(scores):.1f}/{max_score:g}"),
        ("최고점", f"{max(scores) if scores else 0:g}/{max_score:g}"),
        ("중앙값", f"{median(scores) if scores else 0:.1f}"),
        ("표준편차", f"{pstdev(scores) if scores else 0:.1f}"),
        (
            "1차 합격",
            (
                f"{pass_count}명 · {pass_rate * 100:.0f}%"
                if has_pass_criterion
                else "기준 미설정"
            ),
        ),
    ]
    for index, (label, value) in enumerate(metrics):
        column = (index % 4) * 2 + 1
        row = 12 + (index // 4) * 2
        overview.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        overview.merge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1)
        overview.cell(row, column, label).font = Font(size=9, bold=True, color=_MUTED)
        overview.cell(row + 1, column, value).font = Font(size=12, bold=True, color=_NAVY)
        for row_index in (row, row + 1):
            for target_column in (column, column + 1):
                cell = overview.cell(row_index, target_column)
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

    _section_heading(overview, 17, "점수 분포 · 만점 대비 구간", last_column=8)
    distribution_header = 18
    _table_header(overview, distribution_header, ["구간", "원점수 범위", "인원", "비율"])
    bands = [(0, 20), (20, 40), (40, 60), (60, 80), (80, 100)]
    for offset, (lower, upper) in enumerate(bands, start=1):
        count = sum(
            1
            for rate in score_rates
            if (rate >= lower if lower == 0 else rate > lower) and rate <= upper
        )
        row = distribution_header + offset
        overview.cell(row, 1, f"{lower}–{upper}%" if lower == 0 else f">{lower}–{upper}%")
        raw_lower = max_score * lower / 100
        raw_upper = max_score * upper / 100
        overview.cell(
            row,
            2,
            f"{raw_lower:.1f}–{raw_upper:.1f}점"
            if lower == 0
            else f">{raw_lower:.1f}–{raw_upper:.1f}점",
        )
        overview.cell(row, 3, count)
        overview.cell(row, 4, count / len(scores) if scores else 0.0)
        overview.cell(row, 4).number_format = "0.0%"
    _style_data_table(overview, start_row=19, end_row=23, end_column=4)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "점수 구간별 인원"
    chart.y_axis.title = "인원"
    chart.y_axis.scaling.min = 0
    chart.y_axis.majorUnit = 1
    chart.x_axis.title = "만점 대비 구간"
    chart.add_data(Reference(overview, min_col=3, min_row=18, max_row=23), titles_from_data=True)
    chart.set_categories(Reference(overview, min_col=1, min_row=19, max_row=23))
    chart.height = 6.6
    chart.width = 11.5
    chart.legend = None
    overview.add_chart(chart, "F18")

    _section_heading(overview, 26, "보충 우선 문항", last_column=8)
    _table_header(overview, 27, ["우선순위", "문항", "정답률", "정답/응시", "권장 수업 행동"])
    weakest = question_stats[:5]
    if weakest:
        for offset, row_data in enumerate(weakest, start=1):
            accuracy = float(row_data.get("accuracy") or 0.0)
            priority, action = _question_action(accuracy)
            row = 27 + offset
            overview.cell(row, 1, priority)
            overview.cell(row, 2, f"{int(row_data['question_number'])}번")
            overview.cell(row, 3, accuracy)
            overview.cell(row, 3).number_format = "0.0%"
            overview.cell(row, 4, f"{int(row_data.get('correct') or 0)}/{int(row_data.get('attempts') or 0)}")
            overview.cell(row, 5, action)
        _style_data_table(overview, start_row=28, end_row=27 + len(weakest), end_column=5)
    else:
        overview.merge_cells(start_row=28, start_column=1, end_row=29, end_column=5)
        overview.cell(28, 1, "문항별 채점 데이터가 쌓이면 정답률이 낮은 문항부터 표시됩니다.")
        overview.cell(28, 1).alignment = Alignment(vertical="center", horizontal="center")
        overview.cell(28, 1).font = Font(color=_MUTED)

    _finalize_sheet(overview, [15, 18, 14, 16, 18, 16, 18, 18], freeze="A5")
    overview.print_area = "A1:H34"

    priority_sheet = workbook.create_sheet("문항 우선순위")
    _set_title(
        priority_sheet,
        f"{exam.title} · 문항 우선순위",
        "정답률이 낮은 순서입니다. 권장 행동은 수업 검토용이며 점수·컷·재시험 정책을 자동 변경하지 않습니다.",
        last_column=8,
    )
    headers = ["우선순위", "문항", "정답률", "정답 수", "응시 수", "평균 점수", "문항 배점", "권장 행동"]
    _table_header(priority_sheet, 5, headers)
    for offset, row_data in enumerate(question_stats, start=1):
        accuracy = float(row_data.get("accuracy") or 0.0)
        priority, action = _question_action(accuracy)
        values = [
            priority,
            int(row_data["question_number"]),
            accuracy,
            int(row_data.get("correct") or 0),
            int(row_data.get("attempts") or 0),
            float(row_data.get("avg_score") or 0.0),
            float(row_data.get("max_score") or 0.0),
            action,
        ]
        for column, value in enumerate(values, start=1):
            priority_sheet.cell(5 + offset, column, value)
        priority_sheet.cell(5 + offset, 3).number_format = "0.0%"
    if question_stats:
        _style_data_table(priority_sheet, start_row=6, end_row=5 + len(question_stats), end_column=8)
        priority_sheet.conditional_formatting.add(
            f"C6:C{5 + len(question_stats)}",
            ColorScaleRule(
                start_type="num", start_value=0, start_color=_RED_SOFT,
                mid_type="num", mid_value=0.5, mid_color=_AMBER_SOFT,
                end_type="num", end_value=1, end_color=_GREEN_SOFT,
            ),
        )
        priority_sheet.auto_filter.ref = f"A5:H{5 + len(question_stats)}"
    _finalize_sheet(priority_sheet, [12, 10, 12, 11, 11, 13, 12, 24], freeze="A6")

    ranked_sheet = workbook.create_sheet("학생별 등수")
    _set_title(
        ranked_sheet,
        f"{exam.title} · 학생별 등수",
        "석차 기준 1차 점수와 서버 등수입니다. 공동 등수는 같은 등수로 표시하고 다음 등수는 인원만큼 건너뜁니다.",
        last_column=13,
    )
    ranked_headers = [
        "등수", "학교", "이름", "강의", "1차 점수", "만점", "득점률", "평균 대비",
        "합격 기준", "판정", "결과 상태", "오답 문항", "사이트 최종 저장",
    ]
    _table_header(ranked_sheet, 5, ranked_headers)
    ranked_rows: list[tuple[tuple[Any, ...], int, list[Any]]] = []
    for candidate in candidates:
        enrollment_id = int(candidate.enrollment_id)
        result = result_by_id.get(enrollment_id)
        analysis_status = result_status(result)
        scored = analysis_status == "DONE"
        rank_info = rankings.get(enrollment_id, {}) if scored else {}
        score = analysis_score(result) if scored else None
        result_max = (
            float(initial_scores[enrollment_id].max_score or max_score)
            if result is not None
            else max_score
        )
        items = initial_items(result)
        wrong_numbers = sorted(
            question_by_id[int(item.question_id)].number
            for item in items
            if not bool(item.is_correct) and int(item.question_id) in question_by_id
        )
        status = {
            "MISSING": "미응시·미채점",
            "NOT_SUBMITTED": "미응시",
            "PROCESSING": "채점 중",
            "PARTIAL": "채점 미확정",
            "FAILED": "채점 실패",
            "DONE": "완료",
        }[analysis_status]
        achievement = achievement_map.get((enrollment_id, int(exam.id)), {})
        if score is None:
            verdict = "-"
        elif not has_pass_criterion:
            verdict = "기준 미설정"
        elif bool(achievement.get("remediated")):
            verdict = "보충 완료"
        else:
            verdict = "합격" if score >= pass_score else "보충 대상"
        values = [
            rank_info.get("rank"),
            _safe_excel_text(candidate.school),
            _safe_excel_text(candidate.student_name),
            _safe_excel_text(candidate.lecture_title),
            score,
            result_max,
            (score / result_max) if score is not None and result_max > 0 else None,
            (score - _mean(scores)) if score is not None and scores else None,
            pass_score,
            verdict,
            status,
            ", ".join(str(number) for number in wrong_numbers),
            (
                timezone.localtime(initial_scores[enrollment_id].recorded_at).strftime("%Y-%m-%d %H:%M")
                if result is not None and initial_scores[enrollment_id].recorded_at is not None
                else ""
            ),
        ]
        sort_key = (
            rank_info.get("rank") is None,
            rank_info.get("rank") or 10**9,
            _safe_excel_text(candidate.student_name),
            enrollment_id,
        )
        ranked_rows.append((sort_key, enrollment_id, values))
    ranked_rows.sort(key=lambda item: item[0])
    for offset, (_, _, values) in enumerate(ranked_rows, start=1):
        for column, value in enumerate(values, start=1):
            ranked_sheet.cell(5 + offset, column, value)
        ranked_sheet.cell(5 + offset, 7).number_format = "0.0%"
        ranked_sheet.cell(5 + offset, 8).number_format = "+0.0;-0.0;0.0"
        verdict_cell = ranked_sheet.cell(5 + offset, 10)
        if verdict_cell.value == "합격":
            verdict_cell.fill = PatternFill("solid", fgColor=_GREEN_SOFT)
            verdict_cell.font = Font(bold=True, color=_GREEN)
        elif verdict_cell.value == "보충 완료":
            verdict_cell.fill = PatternFill("solid", fgColor=_BLUE_SOFT)
            verdict_cell.font = Font(bold=True, color=_BLUE)
        elif verdict_cell.value == "보충 대상":
            verdict_cell.fill = PatternFill("solid", fgColor=_RED_SOFT)
            verdict_cell.font = Font(bold=True, color=_RED)
    _style_data_table(ranked_sheet, start_row=6, end_row=5 + len(ranked_rows), end_column=13)
    ranked_sheet.auto_filter.ref = f"A5:M{5 + len(ranked_rows)}"
    _finalize_sheet(ranked_sheet, [9, 16, 12, 19, 10, 10, 11, 12, 11, 13, 15, 22, 18], freeze="A6")

    answers_sheet = workbook.create_sheet("학생별 답안")
    answer_last_column = max(7 + len(questions), 8)
    _set_title(
        answers_sheet,
        f"{exam.title} · 학생별 답안",
        "1차 결과 기준입니다. 빨강은 오답, 초록은 정답이며 미입력은 복원 가능한 1차 채점 기록이 없는 문항입니다.",
        last_column=answer_last_column,
    )
    answer_headers = ["등수", "학교", "이름", "강의", "1차 점수", "만점", "판정"] + [
        f"{question.number}번" for question in questions
    ]
    _table_header(answers_sheet, 5, answer_headers)
    for offset, (_, enrollment_id, ranked_values) in enumerate(ranked_rows, start=1):
        result = result_by_id.get(enrollment_id)
        item_by_question = {
            int(item.question_id): item for item in initial_items(result)
        }
        base_values = ranked_values[:6] + [ranked_values[9]]
        for column, value in enumerate(base_values, start=1):
            answers_sheet.cell(5 + offset, column, value)
        for question_offset, question in enumerate(questions, start=8):
            item = item_by_question.get(int(question.question_id))
            cell = answers_sheet.cell(5 + offset, question_offset)
            if item is None:
                cell.value = "미입력"
                cell.font = Font(color=_MUTED)
            else:
                cell.value = _safe_excel_text(str(item.answer or "").strip() or "미입력")
                if bool(item.is_correct):
                    cell.fill = PatternFill("solid", fgColor=_GREEN_SOFT)
                    cell.font = Font(color=_GREEN)
                else:
                    cell.fill = PatternFill("solid", fgColor=_RED_SOFT)
                    cell.font = Font(color=_RED)
    _style_data_table(answers_sheet, start_row=6, end_row=5 + len(ranked_rows), end_column=len(answer_headers))
    answers_sheet.auto_filter.ref = f"A5:{get_column_letter(len(answer_headers))}{5 + len(ranked_rows)}"
    _finalize_sheet(
        answers_sheet,
        [9, 16, 12, 19, 10, 10, 13] + [11] * len(questions),
        freeze="H6",
    )

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()
