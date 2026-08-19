from __future__ import annotations

import io
import logging
import re
import zipfile
from dataclasses import dataclass, field
from typing import Any

from django.db import transaction
from django.db.models import Max
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from apps.domains.results.guards.exam_enrollment_guard import (
    validate_exam_enrollment_assigned,
)
from apps.domains.results.guards.score_edit_lease_guard import (
    require_score_edit_scope_available_for_exam,
)
from apps.domains.results.models import ExamAttempt, Result, ResultFact, ResultItem
from apps.domains.results.utils.exam_absence import current_exam_absence_counts
from apps.support.omr.score_adjustment import get_score_adjustment_from_answers
from apps.support.omr.score_shape import get_exam_score_shape
from apps.support.omr.sheet_resolver import resolve_omr_sheet_for_exam
from apps.support.results.admin_exam_dependencies import dispatch_progress_pipeline
from apps.support.results.exam_result_excel_import_dependencies import (
    ResultImportCandidateRecord,
    get_answer_key_answers,
    get_locked_enrollment_for_tenant,
    get_result_import_candidates,
    get_result_import_questions,
)


logger = logging.getLogger(__name__)

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ROWS = 2_000
MAX_COLUMNS = 600
MAX_HEADER_SCAN_ROWS = 30

_NAME_HEADERS = {"이름", "학생명", "성명", "name", "studentname"}
_ENROLLMENT_HEADERS = {
    "수강등록id",
    "수강등록번호",
    "enrollmentid",
    "enrollment",
}
_STUDENT_PHONE_HEADERS = {
    "학생연락처",
    "학생전화번호",
    "학생핸드폰",
    "학생휴대폰",
    "studentphone",
}
_PARENT_PHONE_HEADERS = {
    "부모님연락처",
    "부모연락처",
    "학부모연락처",
    "학부모전화번호",
    "보호자연락처",
    "parentphone",
}
_ABSENCE_HEADERS = {
    "결시",
    "미응시",
    "시험미응시",
    "응시여부",
    "시험응시여부",
    "응시상태",
    "absent",
    "absence",
    "nottaken",
    "notsubmitted",
}
_CORRECT_MARKERS = {"o", "○", "◯", "정답", "맞음", "맞아요", "true", "1", "v", "✓"}
_WRONG_MARKERS = {"x", "×", "✕", ".", "오답", "틀림", "틀렸음", "false"}
_REVIEW_MARKERS = {"0"}
_ABSENCE_MARKERS = {
    "결시",
    "미응시",
    "시험미응시",
    "o",
    "○",
    "◯",
    "x",
    "×",
    "✕",
    "v",
    "✓",
    "y",
    "yes",
    "예",
    "true",
    "1",
    ".",
}
_PRESENT_MARKERS = {"응시", "응시함", "참석", "n", "no", "아니오", "false", "0"}


class ExamResultWorkbookError(ValueError):
    pass


@dataclass(frozen=True)
class QuestionSpec:
    question_id: int
    number: int
    kind: str
    max_score: float


@dataclass(frozen=True)
class CorrectnessMark:
    is_correct: bool
    include_in_wrong_note: bool = False
    earned_score: float | None = None


@dataclass(frozen=True)
class Candidate:
    enrollment_id: int
    student_name: str
    student_phone: str
    parent_phone: str
    school: str
    lecture_id: int | None
    lecture_title: str
    lecture_color: str
    lecture_chip_label: str
    exam_not_submitted_count: int = 0
    is_not_submitted_for_exam: bool = False

    @property
    def lectures_payload(self) -> list[dict[str, Any]]:
        if not self.lecture_title:
            return []
        return [
            {
                "id": self.lecture_id,
                "lecture_name": self.lecture_title,
                "color": self.lecture_color or None,
                "chip_label": self.lecture_chip_label or None,
            }
        ]


@dataclass(frozen=True)
class PlannedRow:
    source_sheet: str
    source_row: int
    candidate: Candidate
    correctness: dict[int, CorrectnessMark]
    correct_count: int
    wrong_question_numbers: tuple[int, ...]
    review_question_numbers: tuple[int, ...]
    total_score: float
    max_score: float
    will_overwrite: bool
    is_not_submitted: bool
    exam_not_submitted_count: int


@dataclass
class ImportPlan:
    exam: Any
    filename: str
    questions: list[QuestionSpec]
    worksheet_names: list[str] = field(default_factory=list)
    rows: list[PlannedRow] = field(default_factory=list)
    errors: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def can_apply(self) -> bool:
        return bool(self.rows) and not self.errors

    def as_payload(self, *, applied: bool = False) -> dict[str, Any]:
        overwrite_count = sum(1 for row in self.rows if row.will_overwrite)
        not_submitted_count = sum(1 for row in self.rows if row.is_not_submitted)
        return {
            "ok": self.can_apply,
            "applied": bool(applied),
            "exam_id": int(self.exam.id),
            "exam_title": str(self.exam.title or ""),
            "filename": self.filename,
            "worksheet_names": list(self.worksheet_names),
            "question_count": len(self.questions),
            "matched_count": len(self.rows),
            "overwrite_count": overwrite_count,
            "not_submitted_count": not_submitted_count,
            "errors": self.errors,
            "warnings": self.warnings,
            "rows": [
                {
                    "sheet": row.source_sheet,
                    "row": row.source_row,
                    "enrollment_id": row.candidate.enrollment_id,
                    "student_name": row.candidate.student_name,
                    "lectures": row.candidate.lectures_payload,
                    "correct_count": row.correct_count,
                    "wrong_count": len(row.wrong_question_numbers),
                    "wrong_questions": list(row.wrong_question_numbers),
                    "review_count": len(row.review_question_numbers),
                    "review_questions": list(row.review_question_numbers),
                    "total_score": row.total_score,
                    "max_score": row.max_score,
                    "will_overwrite": row.will_overwrite,
                    "is_not_submitted": row.is_not_submitted,
                    "exam_not_submitted_count": row.exam_not_submitted_count,
                }
                for row in self.rows
            ],
        }


def build_exam_result_template(*, exam: Any, tenant: Any) -> bytes:
    questions = _question_specs(exam=exam, tenant=tenant)
    candidates = _exam_candidates(exam=exam, tenant=tenant)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "시험결과"

    last_column = 7 + len(questions) + 1
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet.cell(1, 1, _safe_excel_text(f"{exam.title} · 문항별 정오 입력"))
    sheet.cell(1, 1).font = Font(size=15, bold=True, color="FFFFFF")
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor="1D4ED8")
    sheet.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28

    guides = [
        "작성 방법: 정답은 빈칸/O, 오답은 X, 맞았지만 복습할 문항은 숫자 0으로 표시하세요.",
        "전 문항이 비어 있으면 응시 여부에서 '응시'(만점) 또는 '결시'를 꼭 선택하세요.",
        "객관식·단답형이 섞여 있어도 문항 번호 기준으로 반영됩니다.",
        "수강등록ID와 학생 정보는 수정하지 마세요. 점수는 업로드 후 자동 계산됩니다.",
        "기존에 쓰던 엑셀도 이름(또는 연락처)과 1, 2, 3… 문항 열이 있으면 업로드할 수 있습니다.",
    ]
    for offset, text in enumerate(guides, start=3):
        sheet.merge_cells(
            start_row=offset,
            start_column=1,
            end_row=offset,
            end_column=last_column,
        )
        sheet.cell(offset, 1, text)
        sheet.cell(offset, 1).font = Font(size=10, color="475569")

    header_row = 8
    headers: list[Any] = [
        "수강등록ID",
        "학교",
        "이름",
        "학부모연락처",
        "학생연락처",
        "강의",
        "응시 여부",
        *[question.number for question in questions],
        "점수(확인용)",
    ]
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    question_by_number = {question.number: question for question in questions}
    sheet.cell(header_row, 7).fill = PatternFill("solid", fgColor="6B7280")
    for column, question in enumerate(questions, start=8):
        cell = sheet.cell(header_row, column)
        cell.fill = PatternFill(
            "solid",
            fgColor="0F766E" if question.kind == "essay" else "2563EB",
        )
        cell.comment = None

    data_validation = DataValidation(
        type="list",
        formula1='"O,X,0"',
        allow_blank=True,
        error="정답은 빈칸/O, 오답은 X, 맞았지만 복습할 문항은 숫자 0으로 입력해 주세요.",
        errorTitle="정오 표시 확인",
    )
    sheet.add_data_validation(data_validation)
    absence_validation = DataValidation(
        type="list",
        formula1='"응시,결시"',
        allow_blank=True,
        error="응시 여부는 비워 두거나 '응시' 또는 '결시'로 선택해 주세요.",
        errorTitle="응시 여부 확인",
    )
    sheet.add_data_validation(absence_validation)
    absence_fill = PatternFill("solid", fgColor="E5E7EB")
    confirmation_fill = PatternFill("solid", fgColor="FEF3C7")

    for row_index, candidate in enumerate(candidates, start=header_row + 1):
        values = [
            candidate.enrollment_id,
            _safe_excel_text(candidate.school),
            _safe_excel_text(candidate.student_name),
            _safe_excel_text(candidate.parent_phone),
            _safe_excel_text(candidate.student_phone),
            _safe_excel_text(candidate.lecture_title),
            "결시" if candidate.is_not_submitted_for_exam else "",
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row_index, column, value)
        name_cell = sheet.cell(row_index, 3)
        if candidate.exam_not_submitted_count > 0:
            name_cell.fill = absence_fill
            name_cell.comment = Comment(
                f"누적 미응시 {candidate.exam_not_submitted_count}회",
                "Academy",
            )
        for question_column in range(8, 8 + len(questions)):
            sheet.cell(row_index, question_column, "")
        score_column = 8 + len(questions)
        score_terms = []
        for question_column, question_number in enumerate(
            [question.number for question in questions],
            start=8,
        ):
            question = question_by_number[question_number]
            letter = sheet.cell(row_index, question_column).column_letter
            score_terms.append(
                f'IF(OR(UPPER({letter}{row_index})="X",{letter}{row_index}="×"),0,{question.max_score})'
            )
        sheet.cell(
            row_index,
            score_column,
            (
                f'=IF($G{row_index}="결시","결시",'
                f'IF(AND($G{row_index}="",'
                f'COUNTA(H{row_index}:{get_column_letter(7 + len(questions))}{row_index})=0),'
                f'"확인 필요",ROUND({"+".join(score_terms) or "0"},1)))'
            ),
        )

    if candidates:
        absence_validation.add(
            f"G{header_row + 1}:G{header_row + len(candidates)}"
        )
        sheet.conditional_formatting.add(
            f"C{header_row + 1}:C{header_row + len(candidates)}",
            FormulaRule(
                formula=[f'$G{header_row + 1}="결시"'],
                fill=absence_fill,
            ),
        )
        sheet.conditional_formatting.add(
            f"G{header_row + 1}:G{header_row + len(candidates)}",
            FormulaRule(
                formula=[
                    f'AND($G{header_row + 1}="",'
                    f'COUNTA($H{header_row + 1}:'
                    f'${get_column_letter(7 + len(questions))}{header_row + 1})=0)'
                ],
                fill=confirmation_fill,
            ),
        )
        first_question_column = 8
        last_question_column = 7 + len(questions)
        data_validation.add(
            f"{sheet.cell(header_row + 1, first_question_column).coordinate}:"
            f"{sheet.cell(header_row + len(candidates), last_question_column).coordinate}"
        )

    widths = [16, 16, 12, 18, 18, 18, 12]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for column in range(8, 8 + len(questions)):
        sheet.column_dimensions[get_column_letter(column)].width = 5
    sheet.column_dimensions[get_column_letter(8 + len(questions))].width = 13

    sheet.freeze_panes = f"H{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{sheet.cell(header_row, last_column).coordinate}"
    sheet.sheet_view.showGridLines = False

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_exam_wrong_note_export(*, exam: Any, tenant: Any) -> bytes:
    """Export the current representative wrong-note snapshot, one row per student."""
    questions = _question_specs(exam=exam, tenant=tenant)
    candidates = _exam_candidates(exam=exam, tenant=tenant)
    if not candidates:
        raise ExamResultWorkbookError("이 시험에 등록된 학생이 없습니다.")

    question_by_id = {
        question.question_id: question
        for question in questions
    }
    results = {
        int(result.enrollment_id): result
        for result in Result.objects.filter(
            target_type="exam",
            target_id=int(exam.id),
            enrollment_id__in=[
                candidate.enrollment_id for candidate in candidates
            ],
        ).prefetch_related("items")
    }

    export_rows: list[list[Any]] = []
    for candidate in candidates:
        result = results.get(candidate.enrollment_id)
        if result is None:
            continue

        wrong_items = []
        review_items = []
        for item in result.items.all():
            question = question_by_id.get(int(item.question_id))
            if question is None:
                continue
            if not bool(item.is_correct):
                wrong_items.append((question, item))
            elif bool(item.include_in_wrong_note):
                review_items.append((question, item))

        if not wrong_items and not review_items:
            continue
        wrong_items.sort(key=lambda pair: pair[0].number)
        review_items.sort(key=lambda pair: pair[0].number)

        export_rows.append(
            [
                candidate.enrollment_id,
                _safe_excel_text(candidate.school),
                _safe_excel_text(candidate.student_name),
                _safe_excel_text(candidate.lecture_title),
                _safe_excel_text(str(exam.title or "")),
                float(result.total_score or 0.0),
                float(result.max_score or exam.max_score or 0.0),
                len(wrong_items),
                _format_question_numbers(wrong_items),
                _format_item_scores(wrong_items),
                _format_student_answers(wrong_items),
                len(review_items),
                _format_question_numbers(review_items),
                len(wrong_items) + len(review_items),
                timezone.localtime(result.updated_at).strftime("%Y-%m-%d %H:%M"),
            ]
        )

    if not export_rows:
        raise ExamResultWorkbookError("내보낼 오답 또는 오답노트 지정 기록이 없습니다.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "학생별 오답"
    headers = [
        "수강등록ID",
        "학교",
        "이름",
        "강의",
        "시험",
        "총점",
        "만점",
        "오답 수",
        "오답 문항",
        "오답별 점수",
        "학생 답안",
        "복습 지정 수",
        "복습 지정 문항",
        "오답노트 총 문항",
        "사이트 최종 저장",
    ]
    last_column = len(headers)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet.cell(1, 1, _safe_excel_text(f"{exam.title} · 학생별 오답 기록"))
    sheet.cell(1, 1).font = Font(size=15, bold=True, color="FFFFFF")
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor="1D4ED8")
    sheet.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28

    guides = [
        "현재 사이트의 대표 성적에서 오답과 정답·복습 지정 문항만 학생별로 모았습니다.",
        "오답과 복습 지정은 서로 분리되며, 문항 번호는 쉼표로 구분됩니다.",
        "이 파일은 조회용입니다. 사이트 성적을 바꾸려면 정오표 또는 엑셀 가져오기를 이용하세요.",
    ]
    for row_index, guide in enumerate(guides, start=3):
        sheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=last_column,
        )
        sheet.cell(row_index, 1, guide)
        sheet.cell(row_index, 1).font = Font(size=10, color="475569")

    header_row = 7
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, values in enumerate(export_rows, start=header_row + 1):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.alignment = Alignment(
                horizontal="center" if column in {1, 6, 7, 8, 12, 14, 15} else "left",
                vertical="center",
                wrap_text=column in {9, 10, 11, 13},
            )
        sheet.row_dimensions[row_index].height = 24

    widths = [16, 16, 12, 18, 24, 10, 10, 10, 22, 30, 36, 12, 22, 16, 18]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(last_column)}{header_row + len(export_rows)}"
    )
    sheet.sheet_view.showGridLines = False

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def plan_exam_result_import(
    *,
    exam: Any,
    tenant: Any,
    filename: str,
    workbook_bytes: bytes,
) -> ImportPlan:
    questions = _question_specs(exam=exam, tenant=tenant)
    plan = ImportPlan(exam=exam, filename=filename, questions=questions)
    candidates = _exam_candidates(exam=exam, tenant=tenant)
    if not candidates:
        plan.errors.append(_error(None, "students", "이 시험에 등록된 학생이 없습니다."))
        return plan

    by_id = {candidate.enrollment_id: candidate for candidate in candidates}
    by_name: dict[str, list[Candidate]] = {}
    by_phone: dict[str, list[Candidate]] = {}
    for candidate in candidates:
        by_name.setdefault(_normalize_name(candidate.student_name), []).append(candidate)
        for phone in {candidate.student_phone, candidate.parent_phone}:
            normalized = _normalize_phone(phone)
            if normalized:
                by_phone.setdefault(normalized, []).append(candidate)

    try:
        worksheet_selections = _select_worksheets_for_import(
            workbook_bytes=workbook_bytes,
            exam_title=str(exam.title or ""),
            questions=questions,
            by_id=by_id,
            by_name=by_name,
            by_phone=by_phone,
        )
    except ExamResultWorkbookError as exc:
        plan.errors.append(_error(None, "file", str(exc)))
        return plan

    plan.worksheet_names = [
        str(worksheet.title)
        for worksheet, _, _ in worksheet_selections
    ]
    expected_numbers = {question.number for question in questions}
    for worksheet, header_row_number, columns in worksheet_selections:
        found_numbers = set(columns["questions"])
        missing = sorted(expected_numbers - found_numbers)
        extra = sorted(found_numbers - expected_numbers)
        if missing:
            plan.errors.append(
                _error(
                    header_row_number,
                    "questions",
                    f"시험 문항 열이 빠져 있습니다: {', '.join(map(str, missing))}번",
                    sheet=str(worksheet.title),
                )
            )
        if extra:
            plan.errors.append(
                _error(
                    header_row_number,
                    "questions",
                    f"이 시험에 없는 문항 열이 있습니다: {', '.join(map(str, extra))}번",
                    sheet=str(worksheet.title),
                )
            )
    if plan.errors:
        return plan

    existing_enrollment_ids = set(
        Result.objects.filter(
            target_type="exam",
            target_id=int(exam.id),
            enrollment_id__in=list(by_id),
        ).values_list("enrollment_id", flat=True)
    )
    current_exam_absence_ids = {
        candidate.enrollment_id
        for candidate in candidates
        if candidate.is_not_submitted_for_exam
    }
    question_by_number = {question.number: question for question in questions}
    used_enrollment_ids: set[int] = set()

    for worksheet, header_row_number, columns in worksheet_selections:
        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=header_row_number + 1,
                max_row=worksheet.max_row,
                max_col=worksheet.max_column,
                values_only=True,
            ),
            start=header_row_number + 1,
        ):
            row_values = tuple(values)
            identity_values = [
                _value_at(row_values, columns.get("enrollment")),
                _value_at(row_values, columns.get("name")),
                _value_at(row_values, columns.get("student_phone")),
                _value_at(row_values, columns.get("parent_phone")),
            ]
            question_values = [
                _value_at(row_values, column_index)
                for column_index in columns["questions"].values()
            ]
            absence_value = _value_at(row_values, columns.get("absence"))
            if not any(
                _has_value(value)
                for value in identity_values + question_values + [absence_value]
            ):
                continue

            candidate, match_error = _match_candidate(
                row_values=row_values,
                columns=columns,
                by_id=by_id,
                by_name=by_name,
                by_phone=by_phone,
            )
            if match_error:
                plan.errors.append(
                    _error(
                        row_number,
                        "student",
                        match_error,
                        sheet=str(worksheet.title),
                    )
                )
                continue
            assert candidate is not None
            if candidate.enrollment_id in used_enrollment_ids:
                plan.errors.append(
                    _error(
                        row_number,
                        "student",
                        "같은 학생이 선택된 엑셀 시트에 두 번 들어 있습니다.",
                        sheet=str(worksheet.title),
                    )
                )
                continue

            is_not_submitted = _parse_absence_marker(absence_value)
            if is_not_submitted is None:
                plan.errors.append(
                    _error(
                        row_number,
                        "absence",
                        "응시 여부는 비워 두거나 '응시' 또는 '결시'로 입력해 주세요.",
                        sheet=str(worksheet.title),
                    )
                )
                continue

            questions_are_blank = not any(_has_value(value) for value in question_values)
            if (
                not is_not_submitted
                and questions_are_blank
                and not _has_value(absence_value)
                and not columns.get("absence_blank_is_present", False)
            ):
                plan.errors.append(
                    _error(
                        row_number,
                        "attendance_confirmation",
                        (
                            "전 문항이 비어 있어 만점과 결시를 구분할 수 없습니다. "
                            "만점이면 응시 여부에서 '응시'를, 미응시면 '결시'를 "
                            "선택해 주세요. 기존 엑셀은 정답 문항 하나를 O로 표시해도 됩니다."
                        ),
                        sheet=str(worksheet.title),
                    )
                )
                continue

            correctness: dict[int, CorrectnessMark] = {}
            if not is_not_submitted:
                marker_error = False
                for question_number, column_index in columns["questions"].items():
                    raw_marker = _value_at(row_values, column_index)
                    parsed_marker = _parse_correctness_marker(raw_marker)
                    if parsed_marker is None:
                        plan.errors.append(
                            _error(
                                row_number,
                                f"question_{question_number}",
                                (
                                    f"{question_number}번은 빈칸/O(정답), X(오답), "
                                    "숫자 0(정답·오답노트 포함)으로 입력해 주세요."
                                ),
                                sheet=str(worksheet.title),
                            )
                        )
                        marker_error = True
                        continue
                    correctness[question_number] = parsed_marker
                if marker_error:
                    continue

            used_enrollment_ids.add(candidate.enrollment_id)
            correct_count = sum(1 for mark in correctness.values() if mark.is_correct)
            wrong_numbers = tuple(
                sorted(
                    number
                    for number, mark in correctness.items()
                    if not mark.is_correct
                )
            )
            review_numbers = tuple(
                sorted(
                    number
                    for number, mark in correctness.items()
                    if mark.is_correct and mark.include_in_wrong_note
                )
            )
            total_score, max_score = _score_row(
                exam=exam,
                questions=questions,
                correctness=correctness,
            )
            if is_not_submitted:
                total_score = 0.0
            projected_absence_count = candidate.exam_not_submitted_count
            was_not_submitted = candidate.enrollment_id in current_exam_absence_ids
            if is_not_submitted and not was_not_submitted:
                projected_absence_count += 1
            elif not is_not_submitted and was_not_submitted:
                projected_absence_count = max(0, projected_absence_count - 1)
            plan.rows.append(
                PlannedRow(
                    source_sheet=str(worksheet.title),
                    source_row=row_number,
                    candidate=candidate,
                    correctness=correctness,
                    correct_count=correct_count,
                    wrong_question_numbers=wrong_numbers,
                    review_question_numbers=review_numbers,
                    total_score=total_score,
                    max_score=max_score,
                    will_overwrite=candidate.enrollment_id in existing_enrollment_ids,
                    is_not_submitted=is_not_submitted,
                    exam_not_submitted_count=projected_absence_count,
                )
            )

    if not plan.rows and not plan.errors:
        plan.errors.append(_error(None, "rows", "반영할 학생 행을 찾지 못했습니다."))
    overwrite_count = sum(1 for row in plan.rows if row.will_overwrite)
    if overwrite_count:
        plan.warnings.append(
            f"기존 결과가 있는 {overwrite_count}명은 이번 엑셀의 문항별 정오로 갱신됩니다."
        )
    not_submitted_count = sum(1 for row in plan.rows if row.is_not_submitted)
    if not_submitted_count:
        plan.warnings.append(
            f"결시 {not_submitted_count}명은 점수·석차·문항 통계에서 제외됩니다."
        )
    return plan


@transaction.atomic
def apply_exam_result_import(*, plan: ImportPlan) -> dict[str, Any]:
    if not plan.can_apply:
        raise ExamResultWorkbookError("오류가 있는 엑셀은 반영할 수 없습니다.")

    exam = plan.exam
    require_score_edit_scope_available_for_exam(
        exam=exam,
        tenant=exam.tenant,
    )
    question_by_number = {question.number: question for question in plan.questions}
    now = timezone.now()

    for planned_row in plan.rows:
        enrollment_id = int(planned_row.candidate.enrollment_id)
        validate_exam_enrollment_assigned(exam, enrollment_id)
        enrollment = get_locked_enrollment_for_tenant(
            enrollment_id=enrollment_id,
            tenant=exam.tenant,
        )
        if enrollment is None:
            raise ExamResultWorkbookError(
                f"{planned_row.source_row}행 학생의 수강 정보를 찾을 수 없습니다."
            )

        result, attempt = _locked_result_and_attempt(
            exam=exam,
            enrollment=enrollment,
            initial_total=planned_row.total_score,
            initial_max=planned_row.max_score,
            is_not_submitted=planned_row.is_not_submitted,
            now=now,
        )
        if attempt.status == "grading":
            raise ExamResultWorkbookError(
                f"{planned_row.source_row}행 학생은 현재 채점 중이라 반영할 수 없습니다."
            )

        if planned_row.is_not_submitted:
            ResultItem.objects.select_for_update().filter(result=result).delete()
            ResultFact.objects.create(
                target_type="exam",
                target_id=int(exam.id),
                enrollment_id=enrollment_id,
                submission_id=0,
                attempt_id=int(attempt.id),
                question_id=0,
                answer="",
                is_correct=False,
                score=0.0,
                max_score=float(planned_row.max_score),
                source="excel_import",
                meta={
                    "excel_import": True,
                    "status": "NOT_SUBMITTED",
                    "filename": plan.filename,
                    "source_row": planned_row.source_row,
                    "imported_at": now.isoformat(),
                },
            )
            result.attempt = attempt
            result.objective_score = 0.0
            result.total_score = 0.0
            result.max_score = float(planned_row.max_score)
            result.submitted_at = now
            result.save(
                update_fields=[
                    "attempt",
                    "objective_score",
                    "total_score",
                    "max_score",
                    "submitted_at",
                    "updated_at",
                ]
            )
            meta = dict(attempt.meta or {}) if isinstance(attempt.meta, dict) else {}
            meta["status"] = "NOT_SUBMITTED"
            meta["total_score"] = 0.0
            meta["max_score"] = float(planned_row.max_score)
            meta["synced_from_result"] = True
            meta["last_excel_import"] = {
                "filename": plan.filename,
                "source_row": planned_row.source_row,
                "imported_at": now.isoformat(),
            }
            attempt.meta = meta
            attempt.status = "done"
            attempt.save(update_fields=["meta", "status", "updated_at"])
            continue

        objective_score = 0.0
        item_total = 0.0
        for question_number, mark in planned_row.correctness.items():
            question = question_by_number[question_number]
            is_correct = mark.is_correct
            include_in_wrong_note = mark.include_in_wrong_note
            earned = (
                float(mark.earned_score)
                if mark.earned_score is not None
                else (question.max_score if is_correct else 0.0)
            )
            item_total += earned
            if question.kind == "choice":
                objective_score += earned

            existing_item = (
                ResultItem.objects.select_for_update()
                .filter(result=result, question_id=question.question_id)
                .first()
            )
            changed = (
                existing_item is None
                or bool(existing_item.is_correct) != bool(is_correct)
                or bool(existing_item.include_in_wrong_note)
                != bool(include_in_wrong_note)
                or abs(float(existing_item.score or 0.0) - float(earned)) > 0.0001
                or abs(float(existing_item.max_score or 0.0) - float(question.max_score)) > 0.0001
            )
            if changed:
                ResultFact.objects.create(
                    target_type="exam",
                    target_id=int(exam.id),
                    enrollment_id=enrollment_id,
                    submission_id=0,
                    attempt_id=int(attempt.id),
                    question_id=question.question_id,
                    answer="",
                    is_correct=is_correct,
                    score=float(earned),
                    max_score=float(question.max_score),
                    source="excel_import",
                    meta={
                        "excel_import": True,
                        "filename": plan.filename,
                        "source_row": planned_row.source_row,
                        "imported_at": now.isoformat(),
                        "include_in_wrong_note": include_in_wrong_note,
                    },
                )
            ResultItem.objects.update_or_create(
                result=result,
                question_id=question.question_id,
                defaults={
                    "answer": "",
                    "is_correct": is_correct,
                    "include_in_wrong_note": include_in_wrong_note,
                    "score": float(earned),
                    "max_score": float(question.max_score),
                    "source": "excel_import",
                },
            )

        objective_adjustment, total_adjustment = _score_adjustments(
            exam=exam,
            questions=plan.questions,
        )
        total_score = round(item_total + total_adjustment, 2)
        objective_score = round(objective_score + objective_adjustment, 2)
        result.attempt = attempt
        result.objective_score = objective_score
        result.total_score = total_score
        result.max_score = float(planned_row.max_score)
        result.submitted_at = now
        result.save(
            update_fields=[
                "attempt",
                "objective_score",
                "total_score",
                "max_score",
                "submitted_at",
                "updated_at",
            ]
        )

        meta = dict(attempt.meta or {}) if isinstance(attempt.meta, dict) else {}
        meta.pop("status", None)
        meta["total_score"] = total_score
        meta["max_score"] = float(planned_row.max_score)
        meta["synced_from_result"] = True
        meta["last_excel_import"] = {
            "filename": plan.filename,
            "source_row": planned_row.source_row,
            "imported_at": now.isoformat(),
        }
        if int(attempt.attempt_index) == 1 and not isinstance(meta.get("initial_snapshot"), dict):
            meta["initial_snapshot"] = {
                "total_score": total_score,
                "max_score": float(planned_row.max_score),
                "submitted_at": now.isoformat(),
                "source": "excel_result_import",
            }
        attempt.meta = meta
        attempt.status = "done"
        attempt.save(update_fields=["meta", "status", "updated_at"])

    exam_id = int(exam.id)

    def _dispatch_progress() -> None:
        try:
            dispatch_progress_pipeline(exam_id=exam_id)
        except Exception:
            logger.exception(
                "progress pipeline dispatch failed after excel result import (exam=%s)",
                exam_id,
            )

    transaction.on_commit(_dispatch_progress)
    return plan.as_payload(applied=True)


def _question_specs(*, exam: Any, tenant: Any) -> list[QuestionSpec]:
    try:
        sheet = resolve_omr_sheet_for_exam(
            tenant=tenant,
            exam_id=int(exam.id),
            requested_sheet_id=None,
        )
    except ValueError as exc:
        raise ExamResultWorkbookError("시험 문항을 먼저 등록해 주세요.") from exc

    questions = get_result_import_questions(sheet=sheet)
    if not questions:
        raise ExamResultWorkbookError("시험 문항을 먼저 등록해 주세요.")

    score_shape = get_exam_score_shape(exam)
    specs = [
        QuestionSpec(
            question_id=question.question_id,
            number=int(question.number),
            kind=str(score_shape.question_kind(question.question_id) or "choice"),
            max_score=float(
                score_shape.question_max_score(question.question_id, question.score)
            ),
        )
        for question in questions
    ]
    if len({question.number for question in specs}) != len(specs):
        raise ExamResultWorkbookError("시험 문항 번호가 중복되어 있습니다.")
    return specs


def _exam_candidates(*, exam: Any, tenant: Any) -> list[Candidate]:
    records = get_result_import_candidates(
        exam_id=int(exam.id),
        tenant=tenant,
    )
    enrollment_ids = [record.enrollment_id for record in records]
    absence_counts = current_exam_absence_counts(
        tenant=tenant,
        enrollment_ids=enrollment_ids,
    )
    current_exam_absence_ids = set(
        ExamAttempt.objects.filter(
            exam_id=int(exam.id),
            enrollment_id__in=enrollment_ids,
            enrollment__tenant=tenant,
            is_representative=True,
            meta__status="NOT_SUBMITTED",
        ).values_list("enrollment_id", flat=True)
    )
    return [
        _candidate_from_record(
            record,
            exam_not_submitted_count=absence_counts.get(record.enrollment_id, 0),
            is_not_submitted_for_exam=record.enrollment_id
            in current_exam_absence_ids,
        )
        for record in records
    ]


def _candidate_from_record(
    record: ResultImportCandidateRecord,
    *,
    exam_not_submitted_count: int,
    is_not_submitted_for_exam: bool,
) -> Candidate:
    return Candidate(
        enrollment_id=record.enrollment_id,
        student_name=record.student_name,
        student_phone=record.student_phone,
        parent_phone=record.parent_phone,
        school=record.school,
        lecture_id=record.lecture_id,
        lecture_title=record.lecture_title,
        lecture_color=record.lecture_color,
        lecture_chip_label=record.lecture_chip_label,
        exam_not_submitted_count=exam_not_submitted_count,
        is_not_submitted_for_exam=is_not_submitted_for_exam,
    )


def _load_worksheets(workbook_bytes: bytes):
    if not workbook_bytes:
        raise ExamResultWorkbookError("비어 있는 파일입니다.")
    if len(workbook_bytes) > MAX_UPLOAD_BYTES:
        raise ExamResultWorkbookError("엑셀 파일은 10MB 이하만 업로드할 수 있습니다.")
    try:
        with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as archive:
            total_size = sum(info.file_size for info in archive.infolist())
            if total_size > MAX_UNCOMPRESSED_BYTES:
                raise ExamResultWorkbookError("압축을 푼 엑셀 파일이 너무 큽니다.")
    except zipfile.BadZipFile as exc:
        raise ExamResultWorkbookError("올바른 .xlsx 파일이 아닙니다.") from exc

    try:
        workbook = load_workbook(
            io.BytesIO(workbook_bytes),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ExamResultWorkbookError("엑셀 파일을 열 수 없습니다.") from exc
    if not workbook.worksheets:
        raise ExamResultWorkbookError("엑셀 시트를 찾을 수 없습니다.")
    for worksheet in workbook.worksheets:
        if worksheet.max_row is None or worksheet.max_column is None:
            worksheet.calculate_dimension(force=True)
        if worksheet.max_row > MAX_ROWS or worksheet.max_column > MAX_COLUMNS:
            raise ExamResultWorkbookError(
                f"'{worksheet.title}' 시트는 2,000행·600열 이하로 작성해 주세요."
            )
    return workbook.worksheets


def _select_worksheets_for_import(
    *,
    workbook_bytes: bytes,
    exam_title: str,
    questions: list[QuestionSpec],
    by_id: dict[int, Candidate],
    by_name: dict[str, list[Candidate]],
    by_phone: dict[str, list[Candidate]],
):
    worksheets = _load_worksheets(workbook_bytes)
    expected_numbers = {question.number for question in questions}
    parsed: list[tuple[Any, int, dict[str, Any]]] = []
    for worksheet in worksheets:
        try:
            header_row_number, columns = _find_header(worksheet, questions)
        except ExamResultWorkbookError:
            continue
        parsed.append((worksheet, header_row_number, columns))

    if not parsed:
        raise ExamResultWorkbookError(
            "이름·연락처(또는 수강등록ID)와 1, 2, 3… 문항 번호가 있는 "
            "시트를 찾지 못했습니다."
        )

    exact = [
        selection
        for selection in parsed
        if set(selection[2]["questions"]) == expected_numbers
    ]
    if not exact:
        return [parsed[0]]
    if len(exact) == 1:
        return exact

    ranked: list[
        tuple[
            int,
            set[int],
            tuple[Any, int, dict[str, Any]],
        ]
    ] = []
    for selection in exact:
        worksheet, header_row_number, columns = selection
        matched_ids = _worksheet_candidate_ids(
            worksheet=worksheet,
            header_row_number=header_row_number,
            columns=columns,
            by_id=by_id,
            by_name=by_name,
            by_phone=by_phone,
        )
        ranked.append(
            (
                _worksheet_title_affinity(
                    exam_title=exam_title,
                    worksheet_title=str(worksheet.title),
                ),
                matched_ids,
                selection,
            )
        )

    title_affinity = max(item[0] for item in ranked)
    if title_affinity > 0:
        selected = [
            item
            for item in ranked
            if item[0] == title_affinity and item[1]
        ]
        if not selected:
            selected = [item for item in ranked if item[0] == title_affinity]
    else:
        matched_count = max(len(item[1]) for item in ranked)
        selected = [
            item
            for item in ranked
            if len(item[1]) == matched_count
        ]

    if len(selected) == 1:
        return [selected[0][2]]

    occupied: set[int] = set()
    for _, matched_ids, _ in selected:
        if not matched_ids or occupied & matched_ids:
            names = ", ".join(
                f"'{item[2][0].title}'"
                for item in selected
            )
            raise ExamResultWorkbookError(
                f"가져올 시트를 하나로 판별할 수 없습니다: {names}. "
                "시험명에 날짜·회차를 넣거나 해당 시트만 남긴 파일을 업로드해 주세요."
            )
        occupied.update(matched_ids)
    return [item[2] for item in selected]


def _worksheet_candidate_ids(
    *,
    worksheet,
    header_row_number: int,
    columns: dict[str, Any],
    by_id: dict[int, Candidate],
    by_name: dict[str, list[Candidate]],
    by_phone: dict[str, list[Candidate]],
) -> set[int]:
    matched_ids: set[int] = set()
    for values in worksheet.iter_rows(
        min_row=header_row_number + 1,
        max_row=worksheet.max_row,
        max_col=worksheet.max_column,
        values_only=True,
    ):
        row_values = tuple(values)
        if not any(
            _has_value(_value_at(row_values, columns.get(key)))
            for key in ("enrollment", "name", "student_phone", "parent_phone")
        ):
            continue
        candidate, match_error = _match_candidate(
            row_values=row_values,
            columns=columns,
            by_id=by_id,
            by_name=by_name,
            by_phone=by_phone,
        )
        if candidate is not None and match_error is None:
            matched_ids.add(candidate.enrollment_id)
    return matched_ids


def _worksheet_title_affinity(*, exam_title: str, worksheet_title: str) -> int:
    normalized_exam = _normalize_header(exam_title)
    normalized_sheet = _normalize_header(worksheet_title)
    if normalized_exam == normalized_sheet:
        return 100
    if normalized_sheet and normalized_sheet in normalized_exam:
        return 90

    exam_numbers = re.findall(r"\d+", str(exam_title or ""))
    sheet_numbers = re.findall(r"\d+", str(worksheet_title or ""))
    affinity = 0
    if len(exam_numbers) >= 2 and len(sheet_numbers) >= 2:
        if exam_numbers[:2] == sheet_numbers[:2]:
            affinity += 20
    if len(exam_numbers) >= 3 and len(sheet_numbers) >= 3:
        if exam_numbers[2] == sheet_numbers[2]:
            affinity += 5
    return affinity


def _find_header(worksheet, questions: list[QuestionSpec]) -> tuple[int, dict[str, Any]]:
    expected_numbers = {question.number for question in questions}
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(worksheet.max_row, MAX_HEADER_SCAN_ROWS),
            max_col=worksheet.max_column,
            values_only=True,
        ),
        start=1,
    ):
        columns: dict[str, Any] = {"questions": {}}
        for index, value in enumerate(row):
            normalized = _normalize_header(value)
            if normalized in _ENROLLMENT_HEADERS and "enrollment" not in columns:
                columns["enrollment"] = index
            elif normalized in _NAME_HEADERS and "name" not in columns:
                columns["name"] = index
            elif normalized in _STUDENT_PHONE_HEADERS and "student_phone" not in columns:
                columns["student_phone"] = index
            elif normalized in _PARENT_PHONE_HEADERS and "parent_phone" not in columns:
                columns["parent_phone"] = index
            elif normalized in _ABSENCE_HEADERS and "absence" not in columns:
                columns["absence"] = index
                columns["absence_blank_is_present"] = normalized in {
                    "결시",
                    "미응시",
                    "시험미응시",
                    "absent",
                    "absence",
                    "nottaken",
                    "notsubmitted",
                }

        selected_questions, duplicate_questions = _select_question_columns(
            row=row,
            expected_numbers=expected_numbers,
        )
        columns["questions"] = selected_questions

        has_identity = any(
            key in columns
            for key in ("enrollment", "name", "student_phone", "parent_phone")
        )
        has_expected_question = bool(expected_numbers & set(columns["questions"]))
        if not has_identity or not has_expected_question:
            continue
        if duplicate_questions:
            duplicated = ", ".join(map(str, sorted(duplicate_questions)))
            raise ExamResultWorkbookError(f"문항 열이 중복되어 있습니다: {duplicated}번")
        return row_number, columns
    raise ExamResultWorkbookError(
        "이름·연락처(또는 수강등록ID)와 1, 2, 3… 문항 번호가 있는 헤더 행을 찾지 못했습니다."
    )


def _select_question_columns(
    *,
    row: tuple[Any, ...],
    expected_numbers: set[int],
) -> tuple[dict[int, int], set[int]]:
    numbered = [
        (index, _question_number_from_header(value))
        for index, value in enumerate(row)
    ]
    runs: list[list[tuple[int, int]]] = []
    current: list[tuple[int, int]] = []
    for index, number in numbered:
        if number is None:
            if current:
                runs.append(current)
                current = []
            continue
        current.append((index, number))
    if current:
        runs.append(current)

    candidates: list[
        tuple[int, int, dict[int, int], set[int]]
    ] = []
    for run in runs:
        mapping: dict[int, int] = {}
        duplicates: set[int] = set()
        for index, number in run:
            if number in mapping:
                duplicates.add(number)
            else:
                mapping[number] = index
        if expected_numbers.issubset(mapping):
            extras = len(set(mapping) - expected_numbers)
            candidates.append((extras, run[0][0], mapping, duplicates))

    if candidates:
        _, _, mapping, duplicates = min(
            candidates,
            key=lambda item: (item[0], item[1]),
        )
        return mapping, duplicates

    mapping: dict[int, int] = {}
    duplicates: set[int] = set()
    for index, number in numbered:
        if number is None:
            continue
        if number in mapping:
            duplicates.add(number)
        else:
            mapping[number] = index
    return mapping, duplicates


def _match_candidate(
    *,
    row_values: tuple[Any, ...],
    columns: dict[str, Any],
    by_id: dict[int, Candidate],
    by_name: dict[str, list[Candidate]],
    by_phone: dict[str, list[Candidate]],
) -> tuple[Candidate | None, str | None]:
    enrollment_raw = _value_at(row_values, columns.get("enrollment"))
    name = str(_value_at(row_values, columns.get("name")) or "").strip()
    normalized_name = _normalize_name(name)
    phone_values = [
        _normalize_phone(_value_at(row_values, columns.get("student_phone"))),
        _normalize_phone(_value_at(row_values, columns.get("parent_phone"))),
    ]
    phones = [phone for phone in phone_values if phone]

    if _has_value(enrollment_raw):
        enrollment_id = _positive_int(enrollment_raw)
        candidate = by_id.get(enrollment_id or 0)
        if candidate is None:
            return None, "수강등록ID가 이 시험의 학생과 일치하지 않습니다."
        if normalized_name and normalized_name != _normalize_name(candidate.student_name):
            return None, "수강등록ID와 학생 이름이 서로 다릅니다."
        return candidate, None

    phone_matches: dict[int, Candidate] = {}
    for phone in phones:
        for candidate in by_phone.get(phone, []):
            phone_matches[candidate.enrollment_id] = candidate
    if phone_matches:
        matches = list(phone_matches.values())
        if normalized_name:
            named_matches = [
                candidate
                for candidate in matches
                if _normalize_name(candidate.student_name) == normalized_name
            ]
            if not named_matches:
                return None, "연락처와 학생 이름이 서로 다릅니다."
            matches = named_matches
        if len(matches) == 1:
            return matches[0], None
        return None, "연락처가 같은 학생이 여러 명입니다. 수강등록ID를 함께 입력해 주세요."

    if phones:
        return None, "이 시험에 등록된 학생과 연락처가 일치하지 않습니다."

    if normalized_name:
        matches = by_name.get(normalized_name, [])
        if len(matches) == 1:
            return matches[0], None
        if len(matches) > 1:
            return None, "이름이 같은 학생이 여러 명입니다. 연락처 또는 수강등록ID가 필요합니다."
    return None, "이 시험에 등록된 학생과 이름·연락처가 일치하지 않습니다."


def _score_row(
    *,
    exam: Any,
    questions: list[QuestionSpec],
    correctness: dict[int, CorrectnessMark],
) -> tuple[float, float]:
    item_total = 0.0
    for question in questions:
        mark = correctness.get(question.number)
        if mark is None:
            continue
        if mark.earned_score is not None:
            item_total += float(mark.earned_score)
        elif mark.is_correct:
            item_total += question.max_score
    _, total_adjustment = _score_adjustments(exam=exam, questions=questions)
    score_shape = get_exam_score_shape(exam)
    calculated_max = sum(question.max_score for question in questions) + total_adjustment
    max_score = float(score_shape.total_max_score or calculated_max or exam.max_score or 0.0)
    return round(item_total + total_adjustment, 2), round(max_score, 2)


def _score_adjustments(*, exam: Any, questions: list[QuestionSpec]) -> tuple[float, float]:
    score_shape = get_exam_score_shape(exam)
    adjustment = get_score_adjustment_from_answers(
        get_answer_key_answers(template_exam_id=score_shape.template_exam_id)
    )
    has_choice = any(question.kind == "choice" for question in questions)
    has_essay = any(question.kind == "essay" for question in questions)
    objective = float(adjustment.objective if has_choice else 0.0)
    total = objective + float(adjustment.subjective if has_essay else 0.0)
    return objective, total


def _locked_result_and_attempt(
    *,
    exam: Any,
    enrollment: Any,
    initial_total: float,
    initial_max: float,
    is_not_submitted: bool,
    now,
) -> tuple[Result, ExamAttempt]:
    result = (
        Result.objects.select_for_update()
        .filter(
            target_type="exam",
            target_id=int(exam.id),
            enrollment_id=int(enrollment.id),
        )
        .first()
    )
    attempt = None
    if result and result.attempt_id:
        attempt = ExamAttempt.objects.select_for_update().filter(id=result.attempt_id).first()
    if attempt is None:
        attempt = (
            ExamAttempt.objects.select_for_update()
            .filter(
                exam_id=int(exam.id),
                enrollment_id=int(enrollment.id),
                is_representative=True,
            )
            .first()
        )
    if attempt is None:
        attempts = ExamAttempt.objects.select_for_update().filter(
            exam_id=int(exam.id),
            enrollment_id=int(enrollment.id),
        )
        last_index = attempts.aggregate(Max("attempt_index")).get("attempt_index__max") or 0
        attempts.filter(is_representative=True).update(is_representative=False)
        initial_meta = {}
        if not is_not_submitted:
            initial_meta["initial_snapshot"] = {
                "total_score": float(initial_total),
                "max_score": float(initial_max),
                "submitted_at": now.isoformat(),
                "source": "excel_result_import",
            }
        attempt = ExamAttempt.objects.create(
            exam_id=int(exam.id),
            enrollment_id=int(enrollment.id),
            submission_id=0,
            attempt_index=int(last_index) + 1,
            is_retake=bool(last_index),
            is_representative=True,
            status="done",
            meta=initial_meta,
        )
    elif not attempt.is_representative:
        ExamAttempt.objects.filter(
            exam_id=int(exam.id),
            enrollment_id=int(enrollment.id),
            is_representative=True,
        ).exclude(id=attempt.id).update(is_representative=False)
        attempt.is_representative = True
        attempt.save(update_fields=["is_representative", "updated_at"])

    if result is None:
        result = Result.objects.create(
            target_type="exam",
            target_id=int(exam.id),
            enrollment=enrollment,
            attempt=attempt,
            total_score=0.0,
            max_score=float(initial_max),
            objective_score=0.0,
        )
    elif result.attempt_id != attempt.id:
        result.attempt = attempt
        result.save(update_fields=["attempt", "updated_at"])
    return result, attempt


def _parse_correctness_marker(value: Any) -> CorrectnessMark | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return CorrectnessMark(is_correct=True)
    if isinstance(value, bool):
        return CorrectnessMark(is_correct=bool(value))
    if isinstance(value, (int, float)) and float(value) in {0.0, 1.0}:
        if float(value) == 0.0:
            return CorrectnessMark(
                is_correct=True,
                include_in_wrong_note=True,
            )
        return CorrectnessMark(is_correct=True)
    normalized = "".join(str(value).strip().lower().split())
    if normalized in _REVIEW_MARKERS:
        return CorrectnessMark(
            is_correct=True,
            include_in_wrong_note=True,
        )
    if normalized in _CORRECT_MARKERS:
        return CorrectnessMark(is_correct=True)
    if normalized in _WRONG_MARKERS:
        return CorrectnessMark(
            is_correct=False,
            include_in_wrong_note=True,
        )
    return None


def _parse_absence_marker(value: Any) -> bool | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return False
    if isinstance(value, bool):
        return bool(value)
    if isinstance(value, (int, float)) and float(value) in {0.0, 1.0}:
        return bool(int(value))
    normalized = "".join(str(value).strip().lower().split())
    if normalized in _ABSENCE_MARKERS:
        return True
    if normalized in _PRESENT_MARKERS:
        return False
    return None


def _question_number_from_header(value: Any) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float) and value.is_integer():
        number = int(value)
        return number if number > 0 else None
    text = str(value).strip().lower()
    match = re.fullmatch(r"(?:q|문항)?\s*0*(\d+)\s*(?:번)?", text)
    if not match:
        return None
    number = int(match.group(1))
    return number if number > 0 else None


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^0-9a-z가-힣]", "", str(value or "").strip().lower())


def _safe_excel_text(value: Any) -> str:
    text = str(value or "")
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _format_question_numbers(items: list[tuple[QuestionSpec, Any]]) -> str:
    return ", ".join(str(question.number) for question, _ in items)


def _format_item_scores(items: list[tuple[QuestionSpec, Any]]) -> str:
    return ", ".join(
        f"{question.number}번 {float(item.score or 0.0):g}/{float(item.max_score or question.max_score or 0.0):g}"
        for question, item in items
    )


def _format_student_answers(items: list[tuple[QuestionSpec, Any]]) -> str:
    return _safe_excel_text(
        ", ".join(
            f"{question.number}번: {str(item.answer or '').strip() or '미입력'}"
            for question, item in items
        )
    )


def _normalize_name(value: Any) -> str:
    return "".join(str(value or "").strip().lower().split())


def _normalize_phone(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    digits = "".join(character for character in str(value) if character.isdigit())
    if len(digits) == 10 and digits.startswith("10"):
        digits = f"0{digits}"
    return digits


def _positive_int(value: Any) -> int | None:
    try:
        number = int(float(value))
    except (TypeError, ValueError):
        return None
    return number if number > 0 else None


def _value_at(values: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index < 0 or index >= len(values):
        return None
    return values[index]


def _has_value(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _error(
    row: int | None,
    field: str,
    message: str,
    *,
    sheet: str | None = None,
) -> dict[str, Any]:
    payload = {"row": row, "field": field, "message": message}
    if sheet:
        payload["sheet"] = sheet
    return payload
