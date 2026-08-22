"""End-to-end contract tests for exam result Excel imports."""

from __future__ import annotations

import io
import re
import zipfile
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIRequestFactory, force_authenticate

from apps.core.models import Tenant, TenantMembership
from apps.domains.enrollment.models import Enrollment, SessionEnrollment
from apps.domains.exams.models import Exam, ExamEnrollment, ExamQuestion, Sheet
from apps.domains.lectures.models import Lecture, Session
from apps.domains.results.guards.score_edit_lease_guard import ScoreEditLeaseConflict
from apps.domains.results.models import (
    ExamAttempt,
    Result,
    ResultFact,
    ResultItem,
    ScoreEditDraft,
)
from apps.domains.results.services.question_stats_service import QuestionStatsService
from apps.domains.results.services.exam_result_excel_import import (
    ExamResultWorkbookError,
    apply_exam_result_import,
    build_exam_result_template,
    build_exam_wrong_note_export,
    plan_exam_result_import,
)
from apps.domains.results.services.exam_analysis_export import (
    build_exam_analysis_export,
)
from apps.domains.results.utils.exam_absence import current_exam_absence_counts
from apps.domains.results.utils.ranking import compute_exam_rankings
from apps.domains.results.views.admin_exam_results_view import AdminExamResultsView
from apps.domains.results.views.admin_exam_summary_view import AdminExamSummaryView
from apps.domains.results.views.question_stats_views import AdminExamQuestionStatsView
from apps.domains.results.views.admin_exam_result_excel_import_view import (
    AdminExamAnalysisExcelExportView,
    AdminExamResultExcelImportView,
    AdminExamResultExcelTemplateView,
    AdminExamWrongNoteExcelExportView,
)
from apps.domains.students.models import Student


User = get_user_model()


def _workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _multi_sheet_workbook_bytes(
    sheets: list[tuple[str, list[list[object]]]],
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets:
        sheet = workbook.create_sheet(title=title)
        for row in rows:
            sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _without_worksheet_dimension(payload: bytes) -> bytes:
    source_stream = io.BytesIO(payload)
    output_stream = io.BytesIO()
    with zipfile.ZipFile(source_stream) as source, zipfile.ZipFile(
        output_stream,
        mode="w",
        compression=zipfile.ZIP_DEFLATED,
    ) as output:
        for info in source.infolist():
            content = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                content = re.sub(
                    rb'<dimension ref="[^"]+"\s*/>',
                    b"",
                    content,
                    count=1,
                )
            output.writestr(info, content)
    return output_stream.getvalue()


class ExamResultExcelImportTests(TestCase):
    def setUp(self):
        self.factory = APIRequestFactory()
        self.tenant = Tenant.objects.create(
            name="Excel Results",
            code="excel-results",
            is_active=True,
        )
        self.admin = User.objects.create_user(
            username="excel-results-admin",
            password="pw1234",
            tenant=self.tenant,
            is_staff=True,
        )
        TenantMembership.ensure_active(
            tenant=self.tenant,
            user=self.admin,
            role="admin",
        )
        self.lecture = Lecture.objects.create(
            tenant=self.tenant,
            title="수학 A",
            name="수학 A",
            subject="MATH",
            color="#2563eb",
            chip_label="수A",
        )
        self.session = Session.objects.create(
            lecture=self.lecture,
            order=1,
            title="1차시",
        )
        self.exam = Exam.objects.create(
            tenant=self.tenant,
            title="혼합형 시험",
            subject="수학",
            exam_type=Exam.ExamType.REGULAR,
            max_score=100,
            pass_score=60,
        )
        self.exam.sessions.add(self.session)
        self.sheet = Sheet.objects.create(
            exam=self.exam,
            name="MAIN",
            total_questions=2,
            choice_count=1,
            essay_count=1,
        )
        self.choice_question = ExamQuestion.objects.create(
            sheet=self.sheet,
            number=1,
            score=40,
        )
        self.short_question = ExamQuestion.objects.create(
            sheet=self.sheet,
            number=2,
            score=60,
        )
        self.enrollment = self._create_enrollment(
            name="김학생",
            username="excel-student",
            ps_number="EX-001",
            phone="01012345678",
            parent_phone="01098765432",
        )
        ExamEnrollment.objects.create(exam=self.exam, enrollment=self.enrollment)

    def _create_enrollment(
        self,
        *,
        name: str,
        username: str,
        ps_number: str,
        phone: str,
        parent_phone: str,
    ) -> Enrollment:
        user = User.objects.create_user(
            username=username,
            password="pw1234",
            tenant=self.tenant,
        )
        student = Student.objects.create(
            tenant=self.tenant,
            user=user,
            name=name,
            ps_number=ps_number,
            omr_code=phone[-8:],
            phone=phone,
            parent_phone=parent_phone,
            school_type="HIGH",
            high_school="테스트고",
        )
        return Enrollment.objects.create(
            tenant=self.tenant,
            lecture=self.lecture,
            student=student,
            status="ACTIVE",
        )

    def _request(self, method: str, path: str, *, data=None):
        request_method = getattr(self.factory, method)
        request = request_method(path, data=data or {}, format="multipart")
        request.tenant = self.tenant
        force_authenticate(request, user=self.admin)
        return request

    def test_template_contains_roster_and_numbered_question_columns(self):
        payload = build_exam_result_template(exam=self.exam, tenant=self.tenant)

        workbook = load_workbook(io.BytesIO(payload), data_only=False)
        sheet = workbook["시험결과"]

        self.assertEqual(sheet.cell(8, 1).value, "수강등록ID")
        self.assertEqual(sheet.cell(8, 3).value, "이름")
        self.assertEqual(sheet.cell(8, 7).value, "응시 여부")
        self.assertEqual(sheet.cell(8, 8).value, 1)
        self.assertEqual(sheet.cell(8, 9).value, 2)
        self.assertEqual(sheet.cell(9, 1).value, self.enrollment.id)
        self.assertEqual(sheet.cell(9, 3).value, "김학생")
        self.assertIn('$G9="결시"', sheet.cell(9, 10).value)
        self.assertIn('"확인 필요"', sheet.cell(9, 10).value)
        self.assertIn(
            '"응시,결시"',
            {validation.formula1 for validation in sheet.data_validations.dataValidation},
        )

    def test_template_keeps_cumulative_absence_name_shading(self):
        previous_exam = Exam.objects.create(
            tenant=self.tenant,
            title="이전 시험",
            subject="수학",
            exam_type=Exam.ExamType.REGULAR,
            max_score=100,
            pass_score=60,
        )
        previous_attempt = ExamAttempt.objects.create(
            exam=previous_exam,
            enrollment=self.enrollment,
            submission_id=0,
            attempt_index=1,
            is_representative=True,
            status="done",
            meta={"status": "NOT_SUBMITTED"},
        )
        Result.objects.create(
            target_type="exam",
            target_id=previous_exam.id,
            enrollment=self.enrollment,
            attempt=previous_attempt,
            total_score=0,
            max_score=100,
            objective_score=0,
        )
        attempt = ExamAttempt.objects.create(
            exam=self.exam,
            enrollment=self.enrollment,
            submission_id=0,
            attempt_index=1,
            is_representative=True,
            status="done",
            meta={"status": "NOT_SUBMITTED"},
        )
        Result.objects.create(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
            attempt=attempt,
            total_score=0,
            max_score=100,
            objective_score=0,
        )

        payload = build_exam_result_template(exam=self.exam, tenant=self.tenant)
        sheet = load_workbook(io.BytesIO(payload), data_only=False)["시험결과"]

        self.assertEqual(sheet.cell(9, 7).value, "결시")
        self.assertEqual(sheet.cell(9, 3).fill.fgColor.rgb, "00E5E7EB")
        self.assertEqual(sheet.cell(9, 3).comment.text, "누적 미응시 2회")

    def test_template_escapes_formula_like_student_text(self):
        self.enrollment.student.name = "=HYPERLINK(\"https://invalid.example\")"
        self.enrollment.student.save(update_fields=["name", "updated_at"])

        payload = build_exam_result_template(exam=self.exam, tenant=self.tenant)
        workbook = load_workbook(io.BytesIO(payload), data_only=False)

        self.assertEqual(
            workbook["시험결과"].cell(9, 3).value,
            "'=HYPERLINK(\"https://invalid.example\")",
        )

    def test_existing_x_only_spreadsheet_is_matched_and_scored(self):
        payload = _workbook_bytes(
            [
                ["현장인원", "", "", ""],
                ["학교", "이름", "부모님연락처", "학생연락처", "출석", "결시", 1, 2, "점수"],
                ["테스트고", "김학생", "010-9876-5432", "010-1234-5678", "", "", "", "x", 40],
            ]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="기존채점표.xlsx",
            workbook_bytes=payload,
        )

        self.assertTrue(plan.can_apply, plan.errors)
        self.assertEqual(len(plan.rows), 1)
        row = plan.rows[0]
        self.assertEqual(row.candidate.enrollment_id, self.enrollment.id)
        self.assertEqual(row.correct_count, 1)
        self.assertEqual(row.wrong_question_numbers, (2,))
        self.assertEqual(row.total_score, 40.0)
        self.assertEqual(row.max_score, 100.0)
        self.assertFalse(row.is_not_submitted)

    def test_numeric_zero_is_correct_but_included_in_wrong_note(self):
        payload = _workbook_bytes(
            [
                ["이름", "응시 여부", 1, 2],
                ["김학생", "응시", 0, "X"],
            ]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="ymath.xlsx",
            workbook_bytes=payload,
        )

        self.assertTrue(plan.can_apply, plan.errors)
        row = plan.rows[0]
        self.assertEqual(row.correct_count, 1)
        self.assertEqual(row.wrong_question_numbers, (2,))
        self.assertEqual(row.review_question_numbers, (1,))
        self.assertEqual(row.total_score, 40.0)
        self.assertEqual(plan.as_payload()["rows"][0]["review_questions"], [1])

        apply_exam_result_import(plan=plan)

        review_item = ResultItem.objects.get(
            result__target_id=self.exam.id,
            result__enrollment=self.enrollment,
            question=self.choice_question,
        )
        wrong_item = ResultItem.objects.get(
            result__target_id=self.exam.id,
            result__enrollment=self.enrollment,
            question=self.short_question,
        )
        self.assertTrue(review_item.is_correct)
        self.assertTrue(review_item.include_in_wrong_note)
        self.assertEqual(float(review_item.score), 40.0)
        self.assertFalse(wrong_item.is_correct)
        self.assertTrue(wrong_item.include_in_wrong_note)

    def test_ymath_period_is_wrong_and_zero_is_review(self):
        payload = _workbook_bytes(
            [
                [
                    "학교",
                    "이름",
                    "부모님연락처",
                    "학생연락처",
                    "출석",
                    "결시",
                    1,
                    2,
                    "점수",
                    "등수",
                    1,
                    2,
                ],
                [
                    "테스트고",
                    "김학생",
                    "010-9876-5432",
                    "010-1234-5678",
                    "",
                    "",
                    ".",
                    0,
                    60,
                    1,
                    "",
                    ".",
                ],
            ]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="Ymath-기존양식.xlsx",
            workbook_bytes=payload,
        )

        self.assertTrue(plan.can_apply, plan.errors)
        row = plan.rows[0]
        self.assertEqual(row.wrong_question_numbers, (1,))
        self.assertEqual(row.review_question_numbers, (2,))
        self.assertEqual(row.correct_count, 1)
        self.assertEqual(row.total_score, 60.0)

    def test_blank_absence_column_confirms_perfect_score_in_ymath_sheet(self):
        payload = _workbook_bytes(
            [
                ["이름", "결시", 1, 2],
                ["김학생", "", "", ""],
            ]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="Ymath-만점.xlsx",
            workbook_bytes=payload,
        )

        self.assertTrue(plan.can_apply, plan.errors)
        row = plan.rows[0]
        self.assertFalse(row.is_not_submitted)
        self.assertEqual(row.correct_count, 2)
        self.assertEqual(row.total_score, 100.0)

    def test_period_in_absence_column_marks_not_submitted(self):
        payload = _workbook_bytes(
            [
                ["이름", "결시", 1, 2],
                ["김학생", ".", "", ""],
            ]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="Ymath-결시.xlsx",
            workbook_bytes=payload,
        )

        self.assertTrue(plan.can_apply, plan.errors)
        self.assertTrue(plan.rows[0].is_not_submitted)
        self.assertEqual(plan.rows[0].total_score, 0.0)

    def test_multi_sheet_workbook_selects_sheet_matching_exam_roster(self):
        payload = _multi_sheet_workbook_bytes(
            [
                (
                    "다른 시험",
                    [
                        ["이름", "결시", 1, 2],
                        ["다른학생", "", ".", ""],
                    ],
                ),
                (
                    "현재 시험",
                    [
                        ["이름", "결시", 1, 2],
                        ["김학생", "", ".", ""],
                    ],
                ),
            ]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="Ymath-통합.xlsx",
            workbook_bytes=payload,
        )

        self.assertTrue(plan.can_apply, plan.errors)
        self.assertEqual(plan.worksheet_names, ["현재 시험"])
        self.assertEqual(plan.rows[0].source_sheet, "현재 시험")
        self.assertEqual(plan.rows[0].wrong_question_numbers, (1,))

    def test_multi_sheet_workbook_rejects_overlapping_ambiguous_sheets(self):
        payload = _multi_sheet_workbook_bytes(
            [
                (
                    "A반",
                    [
                        ["이름", "결시", 1, 2],
                        ["김학생", "", ".", ""],
                    ],
                ),
                (
                    "B반",
                    [
                        ["이름", "결시", 1, 2],
                        ["김학생", "", "", "."],
                    ],
                ),
            ]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="Ymath-모호.xlsx",
            workbook_bytes=payload,
        )

        self.assertFalse(plan.can_apply)
        self.assertEqual(plan.errors[0]["field"], "file")
        self.assertIn("가져올 시트를 하나로 판별할 수 없습니다", plan.errors[0]["message"])

    def test_multi_sheet_workbook_combines_disjoint_cohort_sheets(self):
        second_enrollment = self._create_enrollment(
            name="이학생",
            username="excel-student-second-sheet",
            ps_number="EX-002",
            phone="01011112222",
            parent_phone="01033334444",
        )
        ExamEnrollment.objects.create(
            exam=self.exam,
            enrollment=second_enrollment,
        )
        self.exam.title = "7/23(목) 대수 Remake 복습 Test (2)"
        self.exam.save(update_fields=["title", "updated_at"])
        payload = _multi_sheet_workbook_bytes(
            [
                (
                    "7.23(목) 2회차",
                    [
                        ["이름", "결시", 1, 2],
                        ["김학생", "", ".", ""],
                    ],
                ),
                (
                    "7.23(목) 2회차 민사",
                    [
                        ["이름", "결시", 1, 2],
                        ["이학생", "", "", "."],
                    ],
                ),
                (
                    "7.21(화) 1회차",
                    [
                        ["이름", "결시", 1, 2],
                        ["김학생", "", "", "."],
                    ],
                ),
            ]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="Ymath-통합.xlsx",
            workbook_bytes=payload,
        )

        self.assertTrue(plan.can_apply, plan.errors)
        self.assertEqual(
            plan.worksheet_names,
            ["7.23(목) 2회차", "7.23(목) 2회차 민사"],
        )
        self.assertEqual(
            {row.candidate.enrollment_id for row in plan.rows},
            {self.enrollment.id, second_enrollment.id},
        )

    def test_existing_absence_column_marks_student_not_submitted(self):
        payload = _workbook_bytes(
            [
                ["학교", "이름", "학생연락처", "결시", 1, 2],
                ["테스트고", "김학생", "010-1234-5678", "O", "", ""],
            ]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="결시포함.xlsx",
            workbook_bytes=payload,
        )

        self.assertTrue(plan.can_apply, plan.errors)
        self.assertEqual(plan.as_payload()["not_submitted_count"], 1)
        row = plan.rows[0]
        self.assertTrue(row.is_not_submitted)
        self.assertEqual(row.correct_count, 0)
        self.assertEqual(row.wrong_question_numbers, ())
        self.assertEqual(row.total_score, 0.0)
        self.assertEqual(row.exam_not_submitted_count, 1)

    def test_all_blank_questions_require_attendance_confirmation(self):
        payload = _workbook_bytes(
            [
                ["이름", "응시 여부", 1, 2],
                ["김학생", "", "", ""],
            ]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="전문항공란.xlsx",
            workbook_bytes=payload,
        )

        self.assertFalse(plan.can_apply)
        self.assertEqual(plan.errors[0]["field"], "attendance_confirmation")
        self.assertIn("만점과 결시를 구분할 수 없습니다", plan.errors[0]["message"])
        self.assertFalse(Result.objects.filter(target_id=self.exam.id).exists())

    def test_present_marker_confirms_all_blank_questions_as_perfect_score(self):
        payload = _workbook_bytes(
            [
                ["이름", "응시 여부", 1, 2],
                ["김학생", "응시", "", ""],
            ]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="만점확인.xlsx",
            workbook_bytes=payload,
        )

        self.assertTrue(plan.can_apply, plan.errors)
        row = plan.rows[0]
        self.assertFalse(row.is_not_submitted)
        self.assertEqual(row.correct_count, 2)
        self.assertEqual(row.wrong_question_numbers, ())
        self.assertEqual(row.total_score, 100.0)

    def test_invalid_absence_marker_is_rejected(self):
        payload = _workbook_bytes(
            [
                ["이름", "결시", 1, 2],
                ["김학생", "아마도", "", ""],
            ]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="잘못된결시.xlsx",
            workbook_bytes=payload,
        )

        self.assertFalse(plan.can_apply)
        self.assertEqual(plan.errors[0]["field"], "absence")

    def test_question_stats_keep_distinct_legacy_facts_without_attempts(self):
        second_enrollment = self._create_enrollment(
            name="이학생",
            username="excel-student-2",
            ps_number="EX-002",
            phone="01011112222",
            parent_phone="01033334444",
        )
        for enrollment, is_correct in (
            (self.enrollment, True),
            (second_enrollment, False),
        ):
            ResultFact.objects.create(
                target_type="exam",
                target_id=self.exam.id,
                enrollment=enrollment,
                submission_id=0,
                attempt=None,
                question_id=self.choice_question.id,
                answer="",
                is_correct=is_correct,
                score=40 if is_correct else 0,
                max_score=40,
                source="legacy_import",
            )

        unmapped_question_id = max(
            self.choice_question.id,
            self.short_question.id,
        ) + 100_000
        ResultFact.objects.create(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
            submission_id=0,
            attempt=None,
            question_id=unmapped_question_id,
            answer="",
            is_correct=False,
            score=0,
            max_score=1,
            source="legacy_import",
        )

        stats = QuestionStatsService.per_question_stats(exam_id=self.exam.id)

        self.assertNotIn(
            unmapped_question_id,
            [row["question_id"] for row in stats],
        )
        self.assertEqual(stats[0]["question_number"], 1)
        self.assertEqual(stats[0]["attempts"], 2)
        self.assertEqual(stats[0]["correct"], 1)

    def test_question_stats_empty_attempt_scope_returns_no_rows(self):
        ResultFact.objects.create(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
            submission_id=0,
            attempt=None,
            question_id=self.choice_question.id,
            answer="1",
            is_correct=True,
            score=40,
            max_score=40,
            source="legacy_import",
        )

        self.assertEqual(
            QuestionStatsService.per_question_stats(
                exam_id=self.exam.id,
                attempt_ids=[],
            ),
            [],
        )

    def test_question_stats_endpoint_uses_only_current_finalized_attempt(self):
        old_attempt = ExamAttempt.objects.create(
            exam=self.exam,
            enrollment=self.enrollment,
            attempt_index=1,
            is_representative=False,
            status="done",
        )
        current_attempt = ExamAttempt.objects.create(
            exam=self.exam,
            enrollment=self.enrollment,
            attempt_index=2,
            is_retake=True,
            is_representative=True,
            status="done",
        )
        Result.objects.create(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
            attempt=current_attempt,
            total_score=40,
            max_score=100,
        )
        for attempt, is_correct in (
            (old_attempt, True),
            (current_attempt, False),
        ):
            ResultFact.objects.create(
                target_type="exam",
                target_id=self.exam.id,
                enrollment=self.enrollment,
                submission_id=attempt.id,
                attempt=attempt,
                question_id=self.choice_question.id,
                answer="1",
                is_correct=is_correct,
                score=40 if is_correct else 0,
                max_score=40,
                source="manual_grid",
            )

        response = AdminExamQuestionStatsView.as_view()(
            self._request(
                "get",
                f"/results/admin/exams/{self.exam.id}/questions/",
            ),
            exam_id=self.exam.id,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["attempts"], 1)
        self.assertEqual(response.data[0]["correct"], 0)

    def test_dimensionless_xlsx_is_matched_and_scored(self):
        payload = _without_worksheet_dimension(
            _workbook_bytes(
                [
                    ["이름", "학생연락처", 1, 2],
                    ["김학생", "01012345678", "", "X"],
                ]
            )
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="dimensionless.xlsx",
            workbook_bytes=payload,
        )

        self.assertTrue(plan.can_apply, plan.errors)
        self.assertEqual(plan.rows[0].wrong_question_numbers, (2,))
        self.assertEqual(plan.rows[0].total_score, 40.0)

    def test_existing_phone_only_spreadsheet_is_matched(self):
        payload = _workbook_bytes(
            [["학생연락처", 1, 2], ["010-1234-5678", "", "X"]]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="연락처채점표.xlsx",
            workbook_bytes=payload,
        )

        self.assertTrue(plan.can_apply, plan.errors)
        self.assertEqual(plan.rows[0].candidate.enrollment_id, self.enrollment.id)
        self.assertEqual(plan.rows[0].wrong_question_numbers, (2,))

    def test_apply_persists_choice_and_short_answer_correctness(self):
        payload = _workbook_bytes(
            [
                ["이름", "학생전화번호", 1, 2],
                ["김학생", "01012345678", "O", "X"],
            ]
        )
        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="시험결과.xlsx",
            workbook_bytes=payload,
        )

        response = apply_exam_result_import(plan=plan)

        self.assertTrue(response["applied"])
        result = Result.objects.get(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
        )
        self.assertEqual(float(result.objective_score), 40.0)
        self.assertEqual(float(result.total_score), 40.0)
        self.assertEqual(float(result.max_score), 100.0)
        items = {
            item.question_id: item
            for item in ResultItem.objects.filter(result=result)
        }
        self.assertTrue(items[self.choice_question.id].is_correct)
        self.assertFalse(items[self.short_question.id].is_correct)
        self.assertEqual(items[self.choice_question.id].source, "excel_import")
        self.assertEqual(
            ResultFact.objects.filter(
                target_type="exam",
                target_id=self.exam.id,
                source="excel_import",
            ).count(),
            2,
        )

    def test_apply_absence_uses_not_submitted_contract_and_excludes_stats(self):
        scored_payload = _workbook_bytes(
            [["수강등록ID", "이름", 1, 2], [self.enrollment.id, "김학생", "O", "X"]]
        )
        apply_exam_result_import(
            plan=plan_exam_result_import(
                exam=self.exam,
                tenant=self.tenant,
                filename="점수.xlsx",
                workbook_bytes=scored_payload,
            )
        )
        self.assertEqual(
            ResultItem.objects.filter(result__enrollment=self.enrollment).count(),
            2,
        )

        absence_payload = _workbook_bytes(
            [
                ["수강등록ID", "이름", "결시", 1, 2],
                [self.enrollment.id, "김학생", "결시", "", ""],
            ]
        )
        response = apply_exam_result_import(
            plan=plan_exam_result_import(
                exam=self.exam,
                tenant=self.tenant,
                filename="결시.xlsx",
                workbook_bytes=absence_payload,
            )
        )

        result = Result.objects.get(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
        )
        attempt = ExamAttempt.objects.get(id=result.attempt_id)
        self.assertTrue(response["applied"])
        self.assertEqual(response["not_submitted_count"], 1)
        self.assertEqual((attempt.meta or {}).get("status"), "NOT_SUBMITTED")
        self.assertEqual(float(result.total_score), 0.0)
        self.assertEqual(float(result.objective_score), 0.0)
        self.assertFalse(ResultItem.objects.filter(result=result).exists())
        self.assertTrue(
            ResultFact.objects.filter(
                attempt=attempt,
                question_id=0,
                source="excel_import",
                meta__status="NOT_SUBMITTED",
            ).exists()
        )
        self.assertEqual(QuestionStatsService.per_question_stats(exam_id=self.exam.id), [])
        self.assertEqual(
            current_exam_absence_counts(
                tenant=self.tenant,
                enrollment_ids=[self.enrollment.id],
            ),
            {self.enrollment.id: 1},
        )

        list_request = self._request(
            "get",
            f"/results/admin/exams/{self.exam.id}/results/",
        )
        with patch(
            "apps.domains.results.views.admin_exam_results_view.compute_exam_rankings",
            return_value={
                self.enrollment.id: {
                    "rank": 1,
                    "percentile": 1.0,
                    "cohort_size": 99,
                    "cohort_avg": 100.0,
                }
            },
        ):
            list_response = AdminExamResultsView.as_view()(
                list_request,
                exam_id=self.exam.id,
            )
        self.assertEqual(list_response.status_code, 200)
        result_row = list_response.data["results"][0]
        self.assertIsNone(result_row["final_score"])
        self.assertIsNone(result_row["rank"])
        self.assertIsNone(result_row["percentile"])
        self.assertIsNone(result_row["cohort_size"])
        self.assertIsNone(result_row["cohort_avg"])
        self.assertEqual(
            result_row["exam_not_submitted_count"],
            1,
        )
        self.assertNotIn(
            self.enrollment.id,
            compute_exam_rankings(exam_id=self.exam.id, tenant=self.tenant),
        )

        summary_request = self._request(
            "get",
            f"/results/admin/exams/{self.exam.id}/summary/",
        )
        summary_response = AdminExamSummaryView.as_view()(
            summary_request,
            exam_id=self.exam.id,
        )
        self.assertEqual(summary_response.status_code, 200)
        self.assertEqual(summary_response.data["participant_count"], 1)
        self.assertEqual(summary_response.data["fail_count"], 0)

    def test_scored_reimport_clears_excel_absence(self):
        absence_payload = _workbook_bytes(
            [
                ["수강등록ID", "이름", "미응시", 1, 2],
                [self.enrollment.id, "김학생", "X", "", ""],
            ]
        )
        apply_exam_result_import(
            plan=plan_exam_result_import(
                exam=self.exam,
                tenant=self.tenant,
                filename="결시.xlsx",
                workbook_bytes=absence_payload,
            )
        )

        scored_payload = _workbook_bytes(
            [
                ["수강등록ID", "이름", "결시", 1, 2],
                [self.enrollment.id, "김학생", "", "O", "X"],
            ]
        )
        response = apply_exam_result_import(
            plan=plan_exam_result_import(
                exam=self.exam,
                tenant=self.tenant,
                filename="정정점수.xlsx",
                workbook_bytes=scored_payload,
            )
        )

        result = Result.objects.get(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
        )
        attempt = ExamAttempt.objects.get(id=result.attempt_id)
        self.assertEqual(response["not_submitted_count"], 0)
        self.assertNotIn("status", attempt.meta or {})
        self.assertEqual(float(result.total_score), 40.0)
        self.assertEqual(ResultItem.objects.filter(result=result).count(), 2)
        question_stats = QuestionStatsService.per_question_stats(exam_id=self.exam.id)
        self.assertEqual(
            [
                (
                    row["question_id"],
                    row["question_number"],
                    row["attempts"],
                    row["correct"],
                )
                for row in question_stats
            ],
            [
                (self.choice_question.id, 1, 1, 1),
                (self.short_question.id, 2, 1, 0),
            ],
        )
        self.assertEqual(
            QuestionStatsService.top_n_wrong_questions(
                exam_id=self.exam.id,
                n=1,
            ),
            [
                {
                    "question_id": self.short_question.id,
                    "question_number": 2,
                    "wrong_count": 1,
                },
            ],
        )
        ranking = compute_exam_rankings(exam_id=self.exam.id, tenant=self.tenant)
        self.assertEqual(ranking[self.enrollment.id]["rank"], 1)
        self.assertEqual(ranking[self.enrollment.id]["cohort_avg"], 40.0)
        self.assertEqual(
            current_exam_absence_counts(
                tenant=self.tenant,
                enrollment_ids=[self.enrollment.id],
            ),
            {},
        )

    def test_exam_summary_does_not_classify_unset_pass_score(self):
        self.exam.pass_score = 0
        self.exam.save(update_fields=["pass_score", "updated_at"])
        apply_exam_result_import(
            plan=plan_exam_result_import(
                exam=self.exam,
                tenant=self.tenant,
                filename="unset-pass-score-summary.xlsx",
                workbook_bytes=_workbook_bytes(
                    [
                        ["수강등록ID", "이름", 1, 2],
                        [self.enrollment.id, "김학생", "O", "X"],
                    ]
                ),
            )
        )

        response = AdminExamSummaryView.as_view()(
            self._request(
                "get",
                f"/results/admin/exams/{self.exam.id}/summary/",
            ),
            exam_id=self.exam.id,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["participant_count"], 1)
        self.assertEqual(response.data["pass_count"], 0)
        self.assertEqual(response.data["fail_count"], 0)
        self.assertEqual(response.data["pass_rate"], 0.0)

    def test_exam_summary_excludes_absence_from_pass_rate_denominator(self):
        absent_enrollment = self._create_enrollment(
            name="결시학생",
            username="excel-student-absent-summary",
            ps_number="EX-ABSENT-SUMMARY",
            phone="01055556666",
            parent_phone="01077778888",
        )
        ExamEnrollment.objects.create(
            exam=self.exam,
            enrollment=absent_enrollment,
        )
        apply_exam_result_import(
            plan=plan_exam_result_import(
                exam=self.exam,
                tenant=self.tenant,
                filename="absence-pass-rate.xlsx",
                workbook_bytes=_workbook_bytes(
                    [
                        ["수강등록ID", "이름", "미응시", 1, 2],
                        [self.enrollment.id, "김학생", "", "O", "O"],
                        [absent_enrollment.id, "결시학생", "O", "", ""],
                    ]
                ),
            )
        )

        response = AdminExamSummaryView.as_view()(
            self._request(
                "get",
                f"/results/admin/exams/{self.exam.id}/summary/",
            ),
            exam_id=self.exam.id,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["participant_count"], 2)
        self.assertEqual(response.data["pass_count"], 1)
        self.assertEqual(response.data["fail_count"], 0)
        self.assertEqual(response.data["pass_rate"], 1.0)

    def test_exam_summary_excludes_unfinalized_attempt_from_score_metrics(self):
        attempt = ExamAttempt.objects.create(
            exam=self.exam,
            enrollment=self.enrollment,
            attempt_index=1,
            is_representative=True,
            status="grading",
        )
        Result.objects.create(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
            attempt=attempt,
            total_score=100,
            max_score=100,
        )

        response = AdminExamSummaryView.as_view()(
            self._request(
                "get",
                f"/results/admin/exams/{self.exam.id}/summary/",
            ),
            exam_id=self.exam.id,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["participant_count"], 1)
        self.assertEqual(response.data["avg_score"], 0.0)
        self.assertEqual(response.data["pass_count"], 0)
        self.assertEqual(response.data["fail_count"], 0)

    def test_exam_summary_fails_closed_for_cross_tenant_enrollment_result(self):
        other_tenant = Tenant.objects.create(
            name="Foreign Results",
            code="foreign-results",
            is_active=True,
        )
        Enrollment.objects.filter(pk=self.enrollment.pk).update(tenant=other_tenant)
        Result.objects.create(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
            total_score=100,
            max_score=100,
        )

        response = AdminExamSummaryView.as_view()(
            self._request(
                "get",
                f"/results/admin/exams/{self.exam.id}/summary/",
            ),
            exam_id=self.exam.id,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["participant_count"], 0)
        self.assertEqual(response.data["pass_count"], 0)

    def test_apply_is_blocked_while_manual_score_editor_holds_lease(self):
        payload = _workbook_bytes(
            [
                ["이름", "학생전화번호", 1, 2],
                ["김학생", "01012345678", "O", "X"],
            ]
        )
        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="manual-editor-active.xlsx",
            workbook_bytes=payload,
        )
        ScoreEditDraft.objects.create(
            session=self.session,
            tenant=self.tenant,
            editor_user=self.admin,
            payload={
                "client_id": "score-tab",
                "changes": [
                    {
                        "type": "examTotal",
                        "examId": self.exam.id,
                        "enrollmentId": self.enrollment.id,
                        "value": 50,
                    }
                ],
            },
        )

        with self.assertRaises(ScoreEditLeaseConflict):
            apply_exam_result_import(plan=plan)

        self.assertFalse(
            Result.objects.filter(
                target_type="exam",
                target_id=self.exam.id,
            ).exists()
        )

    def test_reimport_identical_values_does_not_duplicate_question_facts(self):
        payload = _workbook_bytes(
            [["수강등록ID", "이름", 1, 2], [self.enrollment.id, "김학생", "", "X"]]
        )
        first = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="same.xlsx",
            workbook_bytes=payload,
        )
        apply_exam_result_import(plan=first)
        second = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="same.xlsx",
            workbook_bytes=payload,
        )
        apply_exam_result_import(plan=second)

        self.assertEqual(
            ResultFact.objects.filter(
                target_type="exam",
                target_id=self.exam.id,
                source="excel_import",
            ).count(),
            2,
        )
        self.assertEqual(second.as_payload()["overwrite_count"], 1)

    def test_preview_rejects_unknown_marker_without_writes(self):
        payload = _workbook_bytes(
            [["이름", "학생연락처", 1, 2], ["김학생", "01012345678", "정답", "△"]]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="invalid.xlsx",
            workbook_bytes=payload,
        )

        self.assertFalse(plan.can_apply)
        self.assertIn("2번", plan.errors[0]["message"])
        self.assertFalse(Result.objects.filter(target_id=self.exam.id).exists())

    def test_preview_rejects_conflicting_name_and_phone(self):
        payload = _workbook_bytes(
            [["이름", "학생연락처", 1, 2], ["다른학생", "01012345678", "", "X"]]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="conflicting-student.xlsx",
            workbook_bytes=payload,
        )

        self.assertFalse(plan.can_apply)
        self.assertIn("연락처와 학생 이름", plan.errors[0]["message"])
        self.assertFalse(Result.objects.filter(target_id=self.exam.id).exists())

    def test_linked_session_roster_is_used_when_exam_assignment_is_empty(self):
        ExamEnrollment.objects.filter(exam=self.exam).delete()
        SessionEnrollment.objects.create(
            tenant=self.tenant,
            session=self.session,
            enrollment=self.enrollment,
        )
        payload = _workbook_bytes(
            [["이름", 1, 2], ["김학생", "", "X"]]
        )

        plan = plan_exam_result_import(
            exam=self.exam,
            tenant=self.tenant,
            filename="session-roster.xlsx",
            workbook_bytes=payload,
        )
        apply_exam_result_import(plan=plan)

        self.assertTrue(
            ExamEnrollment.objects.filter(
                exam=self.exam,
                enrollment=self.enrollment,
            ).exists()
        )

    def test_template_and_import_endpoints_use_the_same_contract(self):
        template_request = self._request(
            "get",
            f"/results/admin/exams/{self.exam.id}/result-import/template/",
        )
        template_response = AdminExamResultExcelTemplateView.as_view()(
            template_request,
            exam_id=self.exam.id,
        )
        self.assertEqual(template_response.status_code, 200)
        self.assertTrue(bytes(template_response.content).startswith(b"PK"))

        upload = SimpleUploadedFile(
            "results.xlsx",
            _workbook_bytes(
                [["수강등록ID", "이름", 1, 2], [self.enrollment.id, "김학생", "", "X"]]
            ),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        preview_request = self._request(
            "post",
            f"/results/admin/exams/{self.exam.id}/result-import/",
            data={"file": upload},
        )
        preview_response = AdminExamResultExcelImportView.as_view()(
            preview_request,
            exam_id=self.exam.id,
        )

        self.assertEqual(preview_response.status_code, 200, preview_response.data)
        self.assertTrue(preview_response.data["ok"])
        self.assertEqual(preview_response.data["matched_count"], 1)

    def test_wrong_note_export_contains_current_site_records(self):
        result = Result.objects.create(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
            total_score=60,
            max_score=100,
        )
        ResultItem.objects.create(
            result=result,
            question=self.choice_question,
            answer="=1+1",
            is_correct=False,
            include_in_wrong_note=True,
            score=0,
            max_score=40,
            source="manual_grid",
        )
        ResultItem.objects.create(
            result=result,
            question=self.short_question,
            answer="42",
            is_correct=True,
            include_in_wrong_note=True,
            score=60,
            max_score=60,
            source="manual_grid",
        )

        payload = build_exam_wrong_note_export(exam=self.exam, tenant=self.tenant)
        workbook = load_workbook(io.BytesIO(payload), data_only=True)
        sheet = workbook["학생별 오답"]

        self.assertEqual(sheet.cell(7, 1).value, "수강등록ID")
        self.assertEqual(sheet.cell(7, 9).value, "오답 문항")
        self.assertEqual(sheet.cell(8, 1).value, self.enrollment.id)
        self.assertEqual(sheet.cell(8, 3).value, "김학생")
        self.assertEqual(sheet.cell(8, 6).value, 60)
        self.assertEqual(sheet.cell(8, 8).value, 1)
        self.assertEqual(sheet.cell(8, 9).value, "1")
        self.assertEqual(sheet.cell(8, 10).value, "1번 0/40")
        self.assertEqual(sheet.cell(8, 11).value, "1번: =1+1")
        self.assertEqual(sheet.cell(8, 12).value, 1)
        self.assertEqual(sheet.cell(8, 13).value, "2")
        self.assertEqual(sheet.cell(8, 14).value, 2)

    def test_analysis_export_combines_briefing_questions_rank_and_answers(self):
        second_enrollment = self._create_enrollment(
            name="박학생",
            username="excel-student-analysis-2",
            ps_number="EX-002",
            phone="01022223333",
            parent_phone="01044445555",
        )
        ExamEnrollment.objects.create(exam=self.exam, enrollment=second_enrollment)
        payload = _workbook_bytes(
            [
                ["수강등록ID", "이름", 1, 2],
                [self.enrollment.id, "김학생", "O", "O"],
                [second_enrollment.id, "박학생", "O", "X"],
            ]
        )
        apply_exam_result_import(
            plan=plan_exam_result_import(
                exam=self.exam,
                tenant=self.tenant,
                filename="analysis-source.xlsx",
                workbook_bytes=payload,
            )
        )
        top_result = Result.objects.get(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
        )
        top_item = top_result.items.get(question=self.choice_question)
        top_item.answer = "=1+1"
        top_item.save(update_fields=["answer", "updated_at"])

        exported = build_exam_analysis_export(exam=self.exam, tenant=self.tenant)
        workbook = load_workbook(io.BytesIO(exported), data_only=True)

        self.assertEqual(
            workbook.sheetnames,
            ["수업 브리핑", "문항 우선순위", "학생별 등수", "학생별 답안"],
        )
        briefing = workbook["수업 브리핑"]
        self.assertIn("수업 분석 리포트", briefing.cell(1, 1).value)
        self.assertEqual(briefing.cell(7, 1).value, "표본 확인 후 판단")
        self.assertEqual(briefing.cell(13, 1).value, "2명")
        self.assertEqual(briefing.cell(13, 3).value, "0명")

        questions = workbook["문항 우선순위"]
        self.assertEqual(questions.cell(5, 1).value, "우선순위")
        self.assertEqual(questions.cell(6, 2).value, 2)
        self.assertEqual(questions.cell(6, 3).value, 0.5)

        ranked = workbook["학생별 등수"]
        self.assertEqual(ranked.cell(6, 1).value, 1)
        self.assertEqual(ranked.cell(6, 3).value, "김학생")
        self.assertEqual(ranked.cell(7, 1).value, 2)
        self.assertEqual(ranked.cell(7, 3).value, "박학생")
        self.assertEqual(ranked.cell(7, 10).value, "보충 대상")

        answers = workbook["학생별 답안"]
        self.assertEqual(answers.cell(5, 8).value, "1번")
        self.assertEqual(answers.cell(6, 8).value, "'=1+1")

    def test_analysis_export_keeps_unset_pass_score_unclassified(self):
        self.exam.pass_score = 0
        self.exam.save(update_fields=["pass_score", "updated_at"])
        apply_exam_result_import(
            plan=plan_exam_result_import(
                exam=self.exam,
                tenant=self.tenant,
                filename="unset-pass-score.xlsx",
                workbook_bytes=_workbook_bytes(
                    [
                        ["수강등록ID", "이름", 1, 2],
                        [self.enrollment.id, "김학생", "O", "X"],
                    ]
                ),
            )
        )

        exported = build_exam_analysis_export(exam=self.exam, tenant=self.tenant)
        workbook = load_workbook(io.BytesIO(exported), data_only=True)

        briefing = workbook["수업 브리핑"]
        self.assertEqual(briefing.cell(7, 4).value, "합격 기준 설정 필요")
        self.assertEqual(briefing.cell(15, 7).value, "기준 미설정")
        ranked = workbook["학생별 등수"]
        self.assertEqual(ranked.cell(6, 10).value, "기준 미설정")

    def test_analysis_export_keeps_first_attempt_score_after_retake_result_changes(self):
        apply_exam_result_import(
            plan=plan_exam_result_import(
                exam=self.exam,
                tenant=self.tenant,
                filename="first-attempt-score.xlsx",
                workbook_bytes=_workbook_bytes(
                    [
                        ["수강등록ID", "이름", 1, 2],
                        [self.enrollment.id, "김학생", "O", "O"],
                    ]
                ),
            )
        )
        result = Result.objects.get(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
        )
        self.assertEqual(result.attempt.meta["initial_snapshot"]["total_score"], 100.0)
        result.total_score = 40
        result.save(update_fields=["total_score", "updated_at"])

        exported = build_exam_analysis_export(exam=self.exam, tenant=self.tenant)
        workbook = load_workbook(io.BytesIO(exported), data_only=True)

        ranked = workbook["학생별 등수"]
        answers = workbook["학생별 답안"]
        self.assertEqual(ranked.cell(6, 5).value, 100)
        self.assertEqual(answers.cell(6, 5).value, 100)

    def test_analysis_export_marks_manual_remediation_as_completed(self):
        apply_exam_result_import(
            plan=plan_exam_result_import(
                exam=self.exam,
                tenant=self.tenant,
                filename="remediated.xlsx",
                workbook_bytes=_workbook_bytes(
                    [
                        ["수강등록ID", "이름", 1, 2],
                        [self.enrollment.id, "김학생", "O", "X"],
                    ]
                ),
            )
        )
        achievement = {
            (self.enrollment.id, self.exam.id): {
                "is_pass": False,
                "remediated": True,
                "final_pass": True,
                "clinic_retake": None,
                "achievement": "REMEDIATED",
                "is_provisional": False,
                "meta_status": None,
            }
        }

        with patch(
            "apps.domains.results.services.exam_analysis_export.compute_exam_achievement_bulk",
            return_value=achievement,
        ):
            exported = build_exam_analysis_export(exam=self.exam, tenant=self.tenant)
        workbook = load_workbook(io.BytesIO(exported), data_only=True)

        ranked = workbook["학생별 등수"]
        self.assertEqual(ranked.cell(6, 10).value, "보충 완료")

    def test_analysis_export_endpoint_returns_private_xlsx(self):
        apply_exam_result_import(
            plan=plan_exam_result_import(
                exam=self.exam,
                tenant=self.tenant,
                filename="analysis-source.xlsx",
                workbook_bytes=_workbook_bytes(
                    [
                        ["수강등록ID", "이름", 1, 2],
                        [self.enrollment.id, "김학생", "O", "X"],
                    ]
                ),
            )
        )
        request = self._request(
            "get",
            f"/results/admin/exams/{self.exam.id}/analysis-export/",
        )

        response = AdminExamAnalysisExcelExportView.as_view()(
            request,
            exam_id=self.exam.id,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(bytes(response.content).startswith(b"PK"))
        self.assertIn("analysis.xlsx", response["Content-Disposition"])
        self.assertEqual(response["Cache-Control"], "private, no-store")

    def test_wrong_note_export_omits_students_without_wrong_notes(self):
        result = Result.objects.create(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
            total_score=100,
            max_score=100,
        )
        ResultItem.objects.create(
            result=result,
            question=self.choice_question,
            is_correct=True,
            include_in_wrong_note=False,
            score=40,
            max_score=40,
            source="manual_grid",
        )

        with self.assertRaisesRegex(
            ExamResultWorkbookError,
            "내보낼 오답 또는 오답노트 지정 기록이 없습니다.",
        ):
            build_exam_wrong_note_export(exam=self.exam, tenant=self.tenant)

    def test_wrong_note_export_endpoint_returns_xlsx(self):
        result = Result.objects.create(
            target_type="exam",
            target_id=self.exam.id,
            enrollment=self.enrollment,
            total_score=0,
            max_score=100,
        )
        ResultItem.objects.create(
            result=result,
            question=self.choice_question,
            is_correct=False,
            include_in_wrong_note=True,
            score=0,
            max_score=40,
            source="manual_grid",
        )
        request = self._request(
            "get",
            f"/results/admin/exams/{self.exam.id}/wrong-note-export/",
        )

        response = AdminExamWrongNoteExcelExportView.as_view()(
            request,
            exam_id=self.exam.id,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(bytes(response.content).startswith(b"PK"))
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("wrong_notes.xlsx", response["Content-Disposition"])
        self.assertEqual(response["Cache-Control"], "private, no-store")

    def test_wrong_note_export_endpoint_explains_empty_records(self):
        request = self._request(
            "get",
            f"/results/admin/exams/{self.exam.id}/wrong-note-export/",
        )

        response = AdminExamWrongNoteExcelExportView.as_view()(
            request,
            exam_id=self.exam.id,
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            str(response.data["detail"]),
            "내보낼 오답 또는 오답노트 지정 기록이 없습니다.",
        )

    def test_other_tenant_exam_is_not_accessible(self):
        other_tenant = Tenant.objects.create(
            name="Other",
            code="excel-results-other",
            is_active=True,
        )
        other_exam = Exam.objects.create(
            tenant=other_tenant,
            title="Other exam",
            exam_type=Exam.ExamType.REGULAR,
        )
        other_lecture = Lecture.objects.create(
            tenant=other_tenant,
            title="Other lecture",
            name="Other lecture",
            subject="MATH",
        )
        other_session = Session.objects.create(
            lecture=other_lecture,
            order=1,
            title="Other session",
        )
        other_exam.sessions.add(other_session)
        request = self._request(
            "get",
            f"/results/admin/exams/{other_exam.id}/result-import/template/",
        )

        response = AdminExamResultExcelTemplateView.as_view()(
            request,
            exam_id=other_exam.id,
        )

        self.assertEqual(response.status_code, 404)

        export_request = self._request(
            "get",
            f"/results/admin/exams/{other_exam.id}/wrong-note-export/",
        )
        export_response = AdminExamWrongNoteExcelExportView.as_view()(
            export_request,
            exam_id=other_exam.id,
        )
        self.assertEqual(export_response.status_code, 404)

        analysis_request = self._request(
            "get",
            f"/results/admin/exams/{other_exam.id}/analysis-export/",
        )
        analysis_response = AdminExamAnalysisExcelExportView.as_view()(
            analysis_request,
            exam_id=other_exam.id,
        )
        self.assertEqual(analysis_response.status_code, 404)

    def test_student_cannot_export_wrong_note_records(self):
        request = self.factory.get(
            f"/results/admin/exams/{self.exam.id}/wrong-note-export/"
        )
        request.tenant = self.tenant
        force_authenticate(request, user=self.enrollment.student.user)

        response = AdminExamWrongNoteExcelExportView.as_view()(
            request,
            exam_id=self.exam.id,
        )

        self.assertEqual(response.status_code, 403)

        analysis_request = self.factory.get(
            f"/results/admin/exams/{self.exam.id}/analysis-export/"
        )
        analysis_request.tenant = self.tenant
        force_authenticate(analysis_request, user=self.enrollment.student.user)
        analysis_response = AdminExamAnalysisExcelExportView.as_view()(
            analysis_request,
            exam_id=self.exam.id,
        )
        self.assertEqual(analysis_response.status_code, 403)
